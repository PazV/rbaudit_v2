#--*-- coding: utf-8 --*--
from flask import Flask, render_template, flash, redirect, url_for, session, request, logging, Blueprint, g, send_file
from wtforms import Form, StringField, TextAreaField, PasswordField, validators
from passlib.hash import sha256_crypt
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import logging
from .login import is_logged_in
import json
import copy
import sys
import traceback
import os
import glob
import re
import shutil
import random
import zipfile
from .db_connection import getDB
db = getDB()
from flask import current_app as app
import app_config as cfg
from . import general_functions
GF = general_functions.GeneralFunctions()
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Alignment, Side, NamedStyle
from openpyxl.cell import Cell
from time import gmtime, strftime, localtime

bp = Blueprint('project', __name__, url_prefix='/project' )

@bp.route('/')
@is_logged_in
def project():
    g=GF.userInfo()
    g.notifications=False
    return render_template('project.html',g=g)

@bp.route('/<project_factor>',methods=['GET','POST'])
@is_logged_in
def goToProject(project_factor):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'project_id':project_id},{'project_factor':project_factor}])
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('project.html',g=g)

@bp.route('/<project_factor>/<form_id>',methods=['GET','POST'])
@is_logged_in
def resolveForm(project_factor,form_id):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'project_id':project_id},{'project_factor':project_factor},{'form_id':form_id}])
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('resolve_form.html',g=g)

@bp.route("/<int:project_factor>/createform/step-1", methods=['GET','POST'])
@is_logged_in
def createformStep1(project_factor):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'project_id':project_id},{'project_factor':project_factor}])
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('createform_step1.html',g=g)

@bp.route("/<int:project_factor>/createform/step-1-import", methods=['GET','POST'])
@is_logged_in
def createformStep1Import(project_factor):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'project_id':project_id},{'project_factor':project_factor}])
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('createform_step1_import.html',g=g)

@bp.route("/<int:project_factor>/createform/step-2/<int:form>", methods=['GET','POST'])
@is_logged_in
def createformStep2(form,project_factor):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'form_id':form},{'project_id':project_id},{'project_factor':project_factor}])
    # g.form_id=form
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('createform_step2.html',g=g)

@bp.route("/<int:project_factor>/createform/clone", methods=['GET','POST'])
@is_logged_in
def createformClone(project_factor):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'project_id':project_id},{'project_factor':project_factor}])
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('clone_form.html',g=g)

@bp.route("/<int:project_factor>/createform/clone-from-another-project", methods=['GET','POST'])
@is_logged_in
def createformCloneAnotherProject(project_factor):
    project_id=int(project_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'project_id':project_id},{'project_factor':project_factor}])
    g.project_factor=project_factor
    has_notifications=db.query("""
        select count(*) from project.notification
        where project_id=%s and user_to=%s and read=False
    """%(project_id,session['user_id'])).dictresult()[0]
    if has_notifications['count']==0:
        g.notifications=False
    else:
        g.notifications=True
    return render_template('clone_form_anotherproject.html',g=g)

@bp.route('/saveProject', methods=['GET','POST'])
@is_logged_in
def saveProject():
    #guardar un proyecto
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_projects'})
                if success:
                    if allowed:
                        if data['project_id']==-1:
                            project_info=copy.deepcopy(data)
                            del project_info['project_id']
                            project_info['created']='now'
                            new_project=db.insert('project.project',project_info)
                            user_1={'user_id':project_info['manager'],'project_id':new_project['project_id']}
                            user_2={'user_id':project_info['partner'],'project_id':new_project['project_id']}
                            db.insert('project.project_users',user_1)
                            db.insert('project.project_users',user_2)

                            if data['include_creator']==True: #dependiendo de la bandera include_creator, se agrega el usuario que creó el proyecto a la lista de usuarios del proyecto
                                user_3={'user_id':project_info['created_by'],'project_id':new_project['project_id']}
                                db.insert('project.project_users',user_3)
                            else:
                                get_ws=db.query("""
                                    select workspace_id from system.user where user_id=%s
                                """%data['user_id']).dictresult()
                                if int(get_ws[0]['workspace_id'])==int(data['workspace_id']):
                                    user_3={'user_id':project_info['created_by'],'project_id':new_project['project_id']}
                                    db.insert('project.project_users',user_3)
                            response['msg_response']='El proyecto ha sido creado.'
                        response['success']=True
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getProjects', methods=['GET','POST'])
@is_logged_in
def getProjects():
    #obtener los proyectos a los que se encuentra agregado el usuario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                projects=db.query("""
                    (select a.project_id, a.name, (a.project_id*%d) as project_factor,to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created
                    from project.project a
                    where a.manager=%s or a.partner=%s
                    union
                    select a.project_id, a.name, (a.project_id*%d) as project_factor,to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created
                    from project.project a, project.project_users b
                    where a.project_id=b.project_id
                    and b.user_id=%s)
                    order by created desc
                """%(int(cfg.project_factor),data['user_id'],data['user_id'],int(cfg.project_factor),data['user_id'])).dictresult()
                # app.logger.info(projects)
                response['success']=True
                response['data']=projects
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route("/saveFormStep1", methods=['GET','POST'])
@is_logged_in
def saveFormStep1():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        if data['form_id']==-1: #formulario nuevo
                            form_info=copy.deepcopy(data)
                            del form_info['form_id']
                            form_info['created_by']=form_info['user_id']
                            form_info['create_date']='now()'
                            form_info['status_id']=1
                            form_info['published']=False
                            form_info['user_last_update']=form_info['user_id']
                            columns_info=form_info['columns_info']
                            columns=[]
                            dict_len=len(columns_info)/2
                            for x in range (1,dict_len+1):
                                column={}
                                column['order']=x
                                column['name']=columns_info['col_%s'%x]
                                column['editable']=columns_info['checkcol_%s'%x]
                                columns.append(column)
                            form_info['columns']=str(columns)

                            new_form=db.insert('project.form',form_info)
                            response['success']=True
                            response['form_id']=new_form['form_id']
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentar de nuevo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route("/<path:subpath>/getUnpublishedForms", methods=['GET','POST'])
@is_logged_in
def getUnpublishedForms(subpath):
    #mostrar formularios que están pendientes por publicar
    #no requiere validación de usuario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                forms=db.query("""
                    select
                        form_id, name
                    from
                        project.form
                    where
                        project_id=%s
                        and status_id in (1,2)
                        order by create_date desc
                """%data['project_id']).dictresult()
                response['data']=forms
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/saveFolder', methods=['GET','POST'])
@is_logged_in
def saveFolder():
    #Guardar nueva carpeta
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_folders'})
                if success:
                    if allowed:
                        if data['mode']=='new':
                            db.insert('project.folder',data)
                            response['msg_response']='La carpeta ha sido agregada.'
                        else:
                            db.query("""
                                update project.folder
                                set name='%s' where folder_id=%s
                            """%(data['name'],data['folder_id']))
                            response['msg_response']='La carpeta ha sido actualizada.'
                        response['success']=True
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getMenu',methods=['GET','POST'])
@is_logged_in
def getMenu():
    #obtener el menú de carpetas
    #no requiere validar permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                folders=db.query("""
                    select folder_id, name from project.folder
                    where project_id=%s and parent_id=-1 order by folder_id
                """%data['project_id']).dictresult()

                final_html=""
                if folders!=[]:
                    for f in folders:
                        fh_node='<input type="checkbox" class="form-check-input tree-menu-checkbox folder-checkbox"><li class="selectable-folder"><a href="#" data-folder="%s">%s</a><ul>'%(f['folder_id'],f['name'])
                        sh_node='</ul></li>'
                        res_html=getMenuNodes(f['folder_id'],fh_node,sh_node,data['project_id'])
                        final_html+=res_html
                response['success']=True
                response['menu']='<ul class="file-tree">%s</ul>'%final_html
                # app.logger.info(response['menu'])

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)


def getMenuNodes(folder_id, fh_html,sh_html,project_id):
    #obtiene los nodos del menú
    #no requiere validar permisos
    try:
        folders=db.query("""
            select folder_id, name
            from project.folder
            where parent_id=%s order by folder_id
        """%folder_id).dictresult()
        forms=db.query("""
            select
                form_id, name
            from project.form where project_id=%s and folder_id=%s
            and status_id>2 order by name
        """%(project_id,folder_id)).dictresult()

        node=""

        if folders==[]:
            if forms==[]:
                return fh_html+sh_html
            else:

                forms_str=""
                for f in forms:
                    project_factor=int(project_id)*int(cfg.project_factor)
                    url='/project/%s/%s'%(project_factor,f['form_id'])
                    forms_str+='<input type="checkbox" class="form-check-input tree-menu-checkbox form-checkbox"><li class="selectable-form"><a href="%s" id="%s">%s</a></li>'%(url,f['form_id'],f['name'])
                return fh_html+forms_str+sh_html
        else:
            forms_str=""
            if forms!=[]:
                for f in forms:
                    project_factor=int(project_id)*int(cfg.project_factor)
                    url='/project/%s/%s'%(project_factor,f['form_id'])
                    forms_str+='<input type="checkbox" class="form-check-input tree-menu-checkbox form-checkbox"><li class="selectable-form"><a href="%s" id="%s">%s</a></li>'%(url,f['form_id'],f['name'])
            current_level=""
            for x in folders:
                fh_new_node='<input type="checkbox" class="form-check-input tree-menu-checkbox folder-checkbox"><li class="selectable-folder"><a href="#" data-folder="%s">%s</a><ul>'%(x['folder_id'],x['name'])
                sh_new_node='</ul></li>'
                res_node=getMenuNodes(x['folder_id'],fh_new_node,sh_new_node,project_id)
                current_level+=res_node

            return fh_html+forms_str+current_level+sh_html
    except:
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))

@bp.route('/deleteFolder', methods=['GET','POST'])
@is_logged_in
def deleteFolder():
    #eliminar carpetas del menú
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_folders'})
                if success:
                    if allowed:
                        funcion
                        #función para validar si la carpeta tiene subcarpetas o archivos
                        # db.query("""
                        #     select count(*) from
                        # """)
                        response['success']=True
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route('/createFormTable', methods=['GET','POST'])
@is_logged_in
def createFormTable():
    #obtiene la tabla del formulario para su configuración
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        page=(int(data['page'])*10)-10
                        form_info=db.query("""
                            select a.form_id,a.project_id,a.columns_number, a.rows, a.columns,a.name,
                            to_char(a.last_updated, 'DD-MM-YYYY HH24:MI:SS') as last_updated,
                            (select u.name from system.user u where u.user_id=a.user_last_update) as user_last_update, b.status
                            from project.form_status b, project.form a
                            where a.form_id=%s
                            and a.status_id=b.status_id
                        """%data['form_id']).dictresult()
                        columns=eval(form_info[0]['columns'])
                        #check if table exists
                        table_name='project_%s_form_%s'%(form_info[0]['project_id'],form_info[0]['form_id'])
                        exists=db.query("""
                            select exists(select 1 from information_schema.tables where table_schema='form' and table_name='%s')
                        """%table_name).dictresult()

                        if exists[0]['exists']==False: #create table

                            columns_str=""
                            for x in columns:
                                columns_str+=" ,col_%s text default ''"%x['order']
                            db.query("""
                                CREATE TABLE form.%s(
                                    entry_id serial not null primary key,
                                    form_id integer not null,
                                    project_id integer not null
                                    %s
                                );
                            """%(table_name,columns_str))
                            for i in range(0,int(form_info[0]['rows'])):
                                db.insert('form.%s'%table_name,{'form_id':form_info[0]['form_id'],'project_id':form_info[0]['project_id']})

                        table_str='<table class="table table-bordered table-responsive-md table-striped text-center" id="grdPrefilledForm"><thead><tr>'
                        for c in columns:
                            table_str+='<th class="text-center">%s</th>'%c['name']
                        table_str+='</tr></thead><tbody>'

                        table_info=db.query("""
                            select * from form.%s order by entry_id
                            offset %s limit 10
                        """%(table_name,page)).dictresult()


                        for t in table_info:
                            keys=sorted(t.iteritems())
                            table_str+='<tr>'
                            for k in keys:
                                if k[0].split('_')[0]=='col':
                                    table_str+='<td class="pt-3-half" contenteditable="true" name="%s" data-entry="%s">%s</td>'%(k[0],t['entry_id'],k[1].decode('utf8'))
                                else:
                                    table_str+='</tr>'
                                    break
                        table_str+='</tbody></table>'

                        if int(form_info[0]['rows'])%10!=0:
                            num_buttons=int(int(form_info[0]['rows'])/10)+1
                        else:
                            num_buttons=int(int(form_info[0]['rows'])/10)
                        buttons='<div class="btn-group btn-group-sm" role="group"><input class="paging-toolbar-number" type="text" readonly id="paging_toolbar_number"/>'
                        for b in range (0,num_buttons):
                            buttons+='<button type="button" class="btn btn-secondary form-paging-toolbar" data-number="%s">%s</button>'%(int(b+1),int(b+1))
                        buttons+='</div>'

                        response['form_name']=form_info[0]['name']
                        response['success']=True
                        response['html']=table_str
                        response['paging_toolbar']=buttons
                        response['status']="(%s)"%form_info[0]['status']
                        response['last_updated']='Actualizado %s por %s'%(form_info[0]['last_updated'],form_info[0]['user_last_update'])
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error la intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)


@bp.route('/savePrefilledForm', methods=['GET','POST'])
@is_logged_in
def savePrefilledForm():
    #guardar información al llenar los datos de configuración de la tabla del formulario
    response={}
    try:
        if request.method=='POST':
            # app.logger.info(request.form)
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        # app.logger.info(data)
                        for x in data['table_data']:
                            # app.logger.info(x)
                            update_list=[]
                            for key,value in x.iteritems():
                                if key.split("_")[0]=='col':
                                    update_list.append("%s=$$%s$$"%(key,value))
                            update_str=','.join(e for e in update_list)

                            query="""
                                update form.project_%s_form_%s set %s where entry_id=%s
                            """%(data['project_id'],data['form_id'],update_str,x['entry_id'])
                            # app.logger.info(query)
                            db.query(query)

                        db.query("""
                            update project.form set status_id=2, user_last_update=%s, last_updated='now' where form_id=%s and project_id=%s
                        """%(data['user_id'],data['form_id'],data['project_id']))

                        response['success']=True
                        response['msg_response']='Los cambios han sido guardados.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getFormRevisionUsers', methods=['GET','POST'])
@is_logged_in
def getFormRevisionUsers():
    #carga los usuarios a seleccionar para ser revisores
    #no requiere validar permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                users=db.query("""
                    select
                        a.user_id,
                        a.name
                    from
                        system.user a,
                        project.project_users b
                    where
                        a.user_id=b.user_id
                    and b.project_id=%s
                    order by a.name
                """%data['project_id']).dictresult()
                response['data']=users
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/publishForm', methods=['GET','POST'])
@is_logged_in
def publishForm():
    #publicar un formulario para que sea resuelto por alguien
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        app.logger.info("user allowed")
                        revisions=[]
                        for k,v in data.iteritems():
                            if k.split('_')[0]=='revision':
                                rev={
                                    'form_id':data['form_id'],
                                    'user_id':v,
                                    'revision_number':k.split("_")[1]
                                }
                                db.insert("project.form_revisions",rev)
                                revisions.append('%s:%s'%(k,v))
                        str_revisions=','.join(e for e in revisions)
                        app.logger.info("do revisions")
                        app.logger.info(str_revisions)
                        app.logger.info("""
                            update project.form
                            set status_id=3,
                            notify_assignee=%s,
                            notify_resolved=%s,
                            resolve_before='%s',
                            assigned_to=%s,
                            revisions='%s'
                            where form_id=%s
                            and project_id=%s
                        """%(data['notify_assignee'],data['notify_resolved'],data['resolve_date'],data['assigned_to'],str_revisions,data['form_id'],data['project_id']))
                        db.query("""
                            update project.form
                            set status_id=3,
                            notify_assignee=%s,
                            notify_resolved=%s,
                            resolve_before='%s',
                            assigned_to=%s,
                            revisions='%s'
                            where form_id=%s
                            and project_id=%s
                        """%(data['notify_assignee'],data['notify_resolved'],data['resolve_date'],data['assigned_to'],str_revisions,data['form_id'],data['project_id']))
                        app.logger.info("updated table")
                        table_name='form.project_%s_form_%s'%(data['project_id'],data['form_id'])
                        exists_rev=db.query("""
                            select column_name from information_schema.columns where table_name='project_%s_form_%s' and column_name='rev_1';
                        """%(data['project_id'],data['form_id'])).dictresult()
                        app.logger.info(exists_rev)
                        if exists_rev==[]:
                            db.query("""
                                alter table %s add rev_1 text default ''
                            """%table_name)
                            app.logger.info("agrega columna revisión")
                        notif_info={'project_id':data['project_id'],'form_id':data['form_id']}
                        # GF.createNotification('publish_form',data['project_id'],data['form_id'])
                        GF.createNotification('publish_form',notif_info)
                        if data['notify_assignee']==True:
                            GF.sendMailNotification({'form_id':data['form_id'],'type':'new_form'})
                        response['success']=True
                        response['msg_response']='El formulario ha sido publicado, puede encontrarlo en el menú del lado izquierdo.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getFormToResolve', methods=['GET','POST'])
@is_logged_in
def getFormToResolve():
    #obtiene la tabla para resolverla
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        page=(int(data['page'])*10)-10
                        form_info=db.query("""
                            select a.form_id,a.project_id,a.columns_number, a.rows, a.columns, a.name, a.assigned_to, a.status_id,
                            to_char(a.last_updated, 'DD-MM-YYYY HH24:MI:SS') as last_updated,
                            (select u.name from system.user u where u.user_id=a.user_last_update) as user_last_update, b.status
                            from  project.form_status b, project.form a
                            where a.form_id=%s
                            and a.status_id=b.status_id
                        """%data['form_id']).dictresult()
                        columns=eval(form_info[0]['columns'])

                        #check if table exists
                        table_name='project_%s_form_%s'%(form_info[0]['project_id'],form_info[0]['form_id'])

                        table_str='<table class="table table-bordered table-responsive-md table-striped text-center table-ns" id="grdFormToResolve"><thead><tr>'
                        editables=[]
                        for c in columns:
                            table_str+='<th class="text-center">%s</th>'%c['name']
                            if c['editable']==True:
                                editables.append('col_%s'%c['order'])
                        rev='Revisión'.decode('utf8')
                        table_str+='<th class="text-center">%s</th>'%rev
                        table_str+='</tr></thead><tbody>'

                        table_info=db.query("""
                            select * from form.%s order by entry_id
                            offset %s limit 10
                        """%(table_name,page)).dictresult()

                        if form_info[0]['status_id']==7:
                            for t in table_info:
                                keys=sorted(t.iteritems())
                                table_str+='<tr>'
                                for k in keys:
                                    if k[0].split('_')[0]=='col':
                                        table_str+='<td class="pt-3-half" contenteditable="false" name="%s" data-entry="%s">%s</td>'%(k[0],t['entry_id'],k[1].decode('utf8'))
                                    else:
                                        table_str+='<td class="pt-3-half" contenteditable="false" name="rev_1" data-entry="%s">%s</td>'%(t['entry_id'],t['rev_1'].decode('utf8'))
                                        table_str+='</tr>'
                                        break
                        else:
                            rev_editable=True #columna de revisión es editable
                            if int(data['user_id'])==form_info[0]['assigned_to']: #en caso de que el usuario que va a ver el formulario sea a quien fue asignado
                                rev_editable=False #no se permite editar columna de revisión
                            for t in table_info:
                                keys=sorted(t.iteritems())
                                table_str+='<tr>'
                                for k in keys:
                                    if k[0].split('_')[0]=='col':
                                        if k[0] in editables:
                                            table_str+='<td class="pt-3-half" contenteditable="true" name="%s" data-entry="%s">%s</td>'%(k[0],t['entry_id'],k[1].decode('utf8'))
                                        else:
                                            table_str+='<td class="pt-3-half" contenteditable="false" name="%s" data-entry="%s">%s</td>'%(k[0],t['entry_id'],k[1].decode('utf8'))
                                    else:
                                        if rev_editable==True:
                                            table_str+='<td class="pt-3-half" contenteditable="true" name="rev_1" data-entry="%s">%s</td>'%(t['entry_id'],t['rev_1'].decode('utf8'))
                                        else:
                                            table_str+='<td class="pt-3-half" contenteditable="false" name="rev_1" data-entry="%s">%s</td>'%(t['entry_id'],t['rev_1'].decode('utf8'))
                                        table_str+='</tr>'
                                        break
                        table_str+='</tbody></table>'

                        if int(form_info[0]['rows'])%10!=0:
                            num_buttons=int(int(form_info[0]['rows'])/10)+1
                        else:
                            num_buttons=int(int(form_info[0]['rows'])/10)
                        buttons='<div class="btn-group btn-group-sm" role="group"><input class="paging-toolbar-number" type="text" readonly id="paging_toolbar_numberTR"/>'
                        for b in range (0,num_buttons):
                            buttons+='<button type="button" class="btn btn-secondary form-paging-toolbar" data-number="%s">%s</button>'%(int(b+1),int(b+1))
                        buttons+='</div>'
                        response['form_name']=form_info[0]['name']
                        response['success']=True
                        response['html']=table_str
                        response['paging_toolbar']=buttons
                        response['status']="(%s)"%form_info[0]['status']
                        response['last_updated']='Actualizado %s por %s'%(form_info[0]['last_updated'],form_info[0]['user_last_update'])
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/checkUserIsAllowed',methods=['GET','POST'])
@is_logged_in
def checkUserIsAllowed():
    #revisa si el usuario tiene permitido ver formulario para resolverlo
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        allowed_users=db.query("""
                            select
                                f.assigned_to,
                                --f.revisions,
                                p.manager,
                                p.partner,
                                p.project_id,
                                f.status_id
                            from
                                project.project p,
                                project.form f
                            where
                                f.form_id=%s
                            and f.project_id=p.project_id
                        """%data['form_id']).dictresult()[0]

                        revisions=db.query("""
                            select user_id from project.form_revisions where form_id=%s
                        """%data['form_id']).dictresult()

                        match=False
                        readonly=False
                        if int(allowed_users['partner'])==int(data['user_id']) or int(allowed_users['manager'])==int(data['user_id']): #si es socio o gerente del proyecto, tiene acceso a editar y/o ver formularios en todo momento
                            match=True
                            readonly=False
                        else:
                            if int(allowed_users['status_id'])<5: #si está en proceso de llenado o configuración del formulario
                                for k,v in allowed_users.iteritems():
                                    if v!='project_id' or v!='status_id': #verifica si es socio, gerente o tiene asignado el formulario
                                        if int(v)==int(data['user_id']):
                                            match=True #si tiene asignado el formulario, tiene permisos para verlo
                                            break
                                if match==False:
                                    for r in revisions:
                                        if int(r['user_id'])==int(data['user_id']):
                                            match=True
                                            break
                                    if match==False: #si no se encuentra entre los usuarios del formulario
                                        #se busca si el usuario tiene permisos para ver todos los formularios
                                        allowed_seeing=db.query("""
                                            select see_all_forms
                                            from system.user where user_id=%s
                                        """%data['user_id']).dictresult()
                                        if allowed_seeing[0]['see_all_forms']==True:
                                            match=True #si tiene permiso, se le permite verlo
                                            readonly=True #pero no puede editarlo


                                    # if k=='revisions':
                                    #     revision_list=v.split(",")
                                    #     for r in revision_list:
                                    #         if int(r.split(":")[1])==int(data['user_id']):
                                    #             match=True #si es revisor, tiene permiso de ver el formulario y editarlo
                                    #             break
                                    #         break
                                    # else:
                                    #     if v!='project_id' or v!='status_id':
                                    #         if int(v)==int(data['user_id']):
                                    #             match=True #si tiene asignado el formulario, tiene permisos para verlo
                                    #             break

                            else: #si se encuentra en la etapa de revisión
                                if int(data['user_id'])==int(allowed_users['assigned_to']): #si es el usuario a quien fue asignado el formulario
                                    match=True #puede verlo
                                    readonly=True #no puede editarlo
                                else: #se busca si está entre los revisores
                                    for r in revisions:
                                        if int(r['user_id'])==int(data['user_id']):
                                            match=True #si está entre los revisores, puede verlo
                                            readonly=False #y puede editarlo
                                    # rev_list=allowed_users['revisions'].split(",") #se busca si está entre los revisores
                                    # app.logger.info(rev_list)
                                    # for r in rev_list:
                                    #     if int(r.split(":")[1])==int(data['user_id']):
                                    #         match=True #si está entre los revisores, puede verlo
                                    #         readonly=False #y puede editarlo
                                    #         break
                                    if match==False: #si no se encuentra entre los usuarios del formulario
                                        #se busca si el usuario tiene permisos para ver todos los formularios
                                        allowed_seeing=db.query("""
                                            select see_all_forms from system.user
                                            where user_id=%s
                                        """%data['user_id']).dictresult()[0]
                                        if allowed_seeing['see_all_forms']==True:
                                            match=True #si tiene permisos, se permite ver el formulario
                                            readonly=True #pero no puede editarlo

                        response['match']=match
                        response['success']=True
                        response['readonly']=readonly
                        if match==False:
                            response['msg_response']='Usted no tiene permisos para acceder a este formulario.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getFormDetails', methods=['GET','POST'])
@is_logged_in
def getFormDetails():
    #muestra los detalles del formulario
    #no requiere validar permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                info=db.query("""
                    select
                        f.name as form_name,
                        (select a.name from system.user a where a.user_id=f.assigned_to) as assigned_to,
                        --f.revisions,
                        to_char(f.create_date,'DD-MM-YYYY HH24:MI:SS') as create_date,
                        (select a.name from system.user a where a.user_id=f.created_by) as created_by,
                        to_char(f.resolve_before,'DD-MM-YYYY HH24:MI:SS') as resolve_before,
                        (select c.status from project.form_status c where c.status_id=f.status_id) as status,
                        f.status_id
                    from
                        project.form f
                    where
                        f.form_id=%s
                """%data['form_id']).dictresult()[0]

                html='<p><b>Nombre: </b>%s<br>'%info['form_name'].decode('utf-8')
                html+='<b>Asignado a: </b>%s<br><b>Resolver antes de: </b>%s<br>'%(info['assigned_to'].decode('utf-8'),info['resolve_before'])

                revisions=db.query("""
                    select b.user_id,
                    (select a.name from system.user a where a.user_id=b.user_id) as user_name,
                    b.revision_number
                    from project.form_revisions b
                    where b.form_id=%s order by b.revision_number asc
                """%data['form_id']).dictresult()
                for r in revisions:
                    rev='Revisión %s'%(r['revision_number'])
                    html+='<b>%s: </b>%s<br>'%(rev.decode('utf-8'),r['user_name'].decode('utf-8'))
                # rev_list=info['revisions'].split(',')
                # for r in rev_list:
                #     rev='Revisión %s'%((r.split(":")[0]).split("_")[1])
                #     user_rev=db.query("""
                #         select name from system.user where user_id=%s
                #     """%r.split(":")[1]).dictresult()[0]
                #     html+='<b>%s: </b>%s<br>'%(rev.decode('utf-8'),user_rev['name'].decode('utf-8'))

                html+='<b>Creado: </b>%s<br>'%info['create_date']
                html+='<b>Creado por: </b>%s<br>'%info['created_by'].decode('utf-8')
                app.logger.info("status: %s"%info['status_id'])
                if info['status_id'] in [5,6]:
                    current_revisor=db.query("""
                        select a.name from system.user a, project.form_revisions b
                        where a.user_id=b.user_id and b.form_id=%s and b.currently_assigned=True
                    """%data['form_id']).dictresult()[0]
                    if info['status_id']==5:
                        html+='<b>Estado actual: </b>%s a %s</p>'%(info['status'].decode('utf-8'),current_revisor['name'].decode('utf-8'))
                    else:
                        html+='<b>Estado actual: </b>%s por %s</p>'%(info['status'].decode('utf-8'),current_revisor['name'].decode('utf-8'))
                else:
                    html+='<b>Estado actual: </b>%s</p>'%info['status'].decode('utf-8')

                response['success']=True
                response['data']=html
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/saveResolvingForm', methods=['GET','POST'])
@is_logged_in
def saveResolvingForm():
    #guardar datos al ir resolviendo el formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        current_status=db.query("""
                            select status_id from project.form where form_id=%s
                        """%data['form_id']).dictresult()[0]
                        for x in data['table_data']:
                            update_list=[]
                            for key,value in x.iteritems():
                                if key.split("_")[0]=='col' or key.split("_")[0]=='rev':
                                    update_list.append("%s=$$%s$$"%(key,value))
                            update_str=','.join(e for e in update_list)
                            query="""
                                update form.project_%s_form_%s set %s where entry_id=%s
                            """%(data['project_id'],data['form_id'],update_str,x['entry_id'])
                            # app.logger.info(query)
                            db.query(query)

                        if int(current_status['status_id'])==3:
                            db.query("""
                                update project.form set status_id=4, user_last_update=%s, last_updated='now' where form_id=%s and project_id=%s
                            """%(data['user_id'],data['form_id'],data['project_id']))
                        else:
                            db.query("""
                                update project.form set user_last_update=%s, last_updated='now' where form_id=%s and project_id=%s
                            """%(data['user_id'],data['form_id'],data['project_id']))

                        response['success']=True
                        response['msg_response']='Los cambios han sido guardados.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    app.logger.info(request.form)
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                app.logger.info(request.form)
                response['msg_response']='Ocurrió un error al intentar guardar la información, favor de revisar el formato del texto a guardar (preferentemente, use texto plano o sin formato).'
        else:
            response['success']=False
            app.logger.info(request.form)
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        app.logger.info(request.form)
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/sendFormToRevision', methods=['GET','POST'])
@is_logged_in
def sendFormToRevision():
    #enviar formulario a revisión
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        db.query("""
                            update project.form
                            set
                                status_id=5,
                                user_last_update=%s,
                                last_updated='now',
                                resolved_date='now'
                            where
                                form_id=%s
                            and project_id=%s
                        """%(data['user_id'],data['form_id'],data['project_id']))
                        db.query("""
                            update project.form_revisions set currently_assigned=True
                            where form_id=%s and revision_number=1
                        """%data['form_id'])
                        notif_info={'project_id':data['project_id'],'form_id':data['form_id']}
                        # GF.createNotification('send_to_revision1',data['project_id'],data['form_id'])
                        GF.createNotification('send_to_revision1',notif_info)

                        form=db.query("""
                            select name as form_name,
                            project_id,
                            (select a.name from system.user a where a.user_id=assigned_to) as assignee_name
                            from project.form
                            where form_id=%s
                        """%data['form_id']).dictresult()[0]
                        revision_email=db.query("""
                            select a.email from system.user a, project.form_revisions b
                            where a.user_id=b.user_id and b.revision_number=1
                            and b.form_id=%s
                        """%data['form_id']).dictresult()[0]['email']
                        form['top_img']=cfg.img_revision_form
                        form['bottom_img']=cfg.img_rb_logo
                        link=os.path.join(cfg.host,'project',str(cfg.project_factor*int(form['project_id'])),str(data['form_id']))
                        form['link']='<a href="%s"> aqu&iacute;</a>'%link

                        template='<p style="text-align: center;"><img src="{top_img}" alt="" width="50" height="50" /></p><p><span style="font-family: Verdana, Geneva, sans-serif; color: rgb(93, 93, 92);">Estimado usuario:</span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Le informamos que {assignee_name} ha terminado de resolver el programa de trabajo {form_name}, y se lo ha enviado para su revisi&oacute;n.</span></span></p><p><span style="font-family: Verdana, Geneva, sans-serif; color: rgb(93, 93, 92);">Para acceder, haga click {link}.</span></p><p><br></p><p><img src="{bottom_img}" alt="" width="250" height="70" /></p>'
                        body=template.format(**form)

                        GF.sendMail('Enviado a revisión',body,revision_email)
                        response['success']=True
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/<path:subpath>/getFormsToCheck', methods=['GET','POST'])
@is_logged_in
def getFormsToCheck(subpath):
    #obtiene los formularios que se encuentran pendientes por revisar
    #no requiere validar permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                forms=db.query("""
                    select
                        form_id, name
                    from
                        project.form
                    where
                        project_id=%s
                        and status_id in (5,6)
                        order by create_date desc
                """%data['project_id']).dictresult()
                managers=db.query("""
                    select manager,partner
                    from project.project
                    where project_id=%s
                """%data['project_id']).dictresult()[0]
                response['success']=True
                if int(managers['manager'])==int(data['user_id']) or int(managers['partner'])==int(data['user_id']):
                    response['data']=forms
                else:
                    showing_forms=[]
                    for f in forms:
                        revisions=db.query("""
                            select user_id from project.form_revisions where form_id=%s
                        """%f['form_id']).dictresult()
                        for r in revisions:
                            if int(r['user_id'])==int(data['user_id']):
                                showing_forms.append(f)
                                break
                    response['data']=showing_forms

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/checkToDoRevision', methods=['GET','POST'])
@is_logged_in
def checkToDoRevision():
    #revisa si tiene permiso para revisar el formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        form=db.query("""
                            select
                                status_id
                                --, revisions
                            from
                                project.form
                            where form_id=%s
                        """%(data['form_id'])).dictresult()
                        response['success']=True
                        if form[0]['status_id'] in (5,6): #revisa el status del formulario para ver si puede ser revisado
                            users=db.query("""
                                select manager, partner
                                from project.project
                                where project_id=%s
                            """%data['project_id']).dictresult()[0]
                            if int(data['user_id'])==int(users['manager']) or int(data['user_id'])==int(users['partner']): #verifica si el usuario es el gerente o socio del proyecto
                                response['allowed']=True
                            else:
                                revisions=db.query("""
                                    select user_id from project.form_revisions where form_id=%s
                                    and currently_assigned=True
                                """%data['form_id']).dictresult()[0]
                                # rev_list=form[0]['revisions'].split(",")
                                # app.logger.info("rev_list")
                                # app.logger.info(rev_list)
                                # for r in rev_list: #verifica si el usuario es alguno de los revisores
                                # for r in revisions: #verifica si el usuario es alguno de los revisores
                                #     # if int(data['user_id'])==int(r.split(":")[1]):
                                #     if int(data['user_id'])==int(r['user_id']):
                                #         app.logger.info("allowed")
                                #         response['allowed']=True
                                #         break
                                if int(revisions['user_id'])==int(data['user_id']):
                                    response['allowed']=True
                                if 'allowed' not in response:
                                    response['allowed']=False
                                    response['msg_response']='Usted no tiene permisos para revisar este formulario.'
                        else:
                            response['msg_response']='El formulario no se encuentra disponible para ser revisado.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/checkAddComment', methods=['GET','POST'])
@is_logged_in
def checkAddComment():
    #verificar si tiene permisos para agregar comentarios al formulario
    #no requiere validación de permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:

                # form_info=db.query("""
                #     select project_id
                #     from project.form
                #     where form_id=%s
                # """%data['form_id']).dictresult()[0]
                user_list=[]
                # revs=form_info['revisions'].split(",")
                revisions=db.query("""
                    select user_id from project.form_revisions where form_id=%s
                """%data['form_id']).dictresult()
                # for r in revs:
                #     user_list.append(int(r.split(":")[1]))
                for r in revisions:
                    user_list.append(int(r['user_id']))
                managers=db.query("""
                    select a.manager,a.partner,b.assigned_to, a.project_id from project.project a, project.form b
                    where a.project_id=b.project_id and b.form_id=%s
                """%data['form_id']).dictresult()[0]
                user_list.append(int(managers['manager']))
                user_list.append(int(managers['partner']))
                #se agregó que usuario que resuelve pueda agregar comentarios
                user_list.append(int(managers['assigned_to']))
                #agregar usuario consultor
                if data['is_consultant']==True:
                    #comprueba que sea consultor
                    check_is_c=db.query("""
                        select * from system.consultants where user_id=%s
                    """%data['user_id']).dictresult()

                    if check_is_c!=[]:
                        #obtiene ws de manager
                        manager_ws=db.query("""
                            select workspace_id from system.user where user_id=%s
                        """%managers['manager']).dictresult()
                        c_ws=check_is_c[0]['workspaces'].split(",")

                        if str(manager_ws[0]['workspace_id']) in c_ws:
                            user_list.append(int(data['user_id']))
                if int(data['user_id']) in user_list:
                    form_status=db.query("""
                        select status_id from project.form where form_id=%s
                    """%data['form_id']).dictresult()[0]
                    if int(form_status['status_id'])!=7:
                        response['access']=True
                    else:
                        response['access']=False
                        response['msg_response']='No es posible agregar comentarios a un formulario cerrado.'
                else:
                    response['access']=False
                    response['msg_response']='No tienes permisos para agregar comentarios a este formulario.'
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/addFormComment', methods=['GET','POST'])
@is_logged_in
def addFormComment():
    #agregar un comentario al formulario
    #no requiere validación de permisos
    response={}
    try:
        if request.method=='POST':

            # valid,data=GF.getDict(request.form,'post')
            valid=True
            data=request.form.to_dict()

            if valid:

                comment={
                    # 'comment':data['comment'].encode('utf-8'),
                    'comment':data['comment'].encode('utf-8').replace("\n","<br>").replace("\r","<br>").replace("•","-").replace("\t"," "),
                    # 'comment':data['comment'].encode('ascii',errors='ignore').replace("\n","<br>").replace("\r","<br>").replace("•","-"),
                    'user_id':data['user_id'],
                    'form_id':data['form_id'],
                    'created':'now'
                }
                new_comm=db.insert('project.form_comments',comment)

                users=db.query("""
                    select a.user_id, (select b.email from system.user b where a.user_id=b.user_id) as email from project.form_revisions a where a.form_id=%s
                    union
                    select c.assigned_to as user_id, (select d.email from system.user d where d.user_id=c.assigned_to) as email from project.form c where c.form_id=%s
                """%(data['form_id'],data['form_id'])).dictresult()

                info={
                    'form_id':data['form_id'],
                    'project_id':data['project_id'],
                    'msg':data['comment'].encode('utf-8'),
                    'user_from':data['user_id']
                }
                for u in users:
                    info['user_to']=u['user_id']
                    # se envía notificación a usuarios correspondientes
                    GF.createNotification('add_comment',info)

                    mail_data={
                        'comment_id':new_comm['comment_id'],
                        'recipient_mail':u['email'],
                        'type':'new_form_comment',
                        'form_id':data['form_id']
                    }
                    # GF.sendMailNotification(mail_data)


                response['success']=True
                response['msg_response']='El comentario ha sido agregado.'

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getFormComments', methods=['GET','POST'])
@is_logged_in
def getFormComments():
    #ver los comentarios que han sido agregados al formulario
    #no requiere validación de permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                comments=db.query("""
                    select
                        to_char(b.created, 'DD-MM-YYYY HH24:MI:SS') as created,
                        b.comment,
                        (select a.name from system.user a where a.user_id=b.user_id) as user
                    from
                        project.form_comments b
                    where
                        b.form_id=%s
                    order by created desc
                """%data['form_id']).dictresult()

                if comments!=[]:
                    for c in comments:
                        c['author_date']='{created}'.format(**c)
                        c['author_name']='{user}'.format(**c)
                        c['author']='Agregado por {user} el {created}.'.format(**c)
                response['data']=comments
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/checkSendToRevision', methods=['GET','POST'])
@is_logged_in
def checkSendToRevision():
    #revisar si tiene permisos para enviar a revisión un formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        check=db.query("""
                            select assigned_to,status_id from project.form
                            where form_id=%s
                        """%data['form_id']).dictresult()[0]
                        if check['status_id']==3 or check['status_id']==4:
                            if int(check['assigned_to'])==int(data['user_id']):
                                response['allowed']=True
                            else:
                                response['allowed']=False
                                response['msg_response']='No tienes permiso para enviar a revisión este formulario.'
                        else:
                            response['allowed']=False
                            response['msg_response']='Este formulario no puede ser enviado a revisión'
                        response['success']=True
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/importNewForm',methods=['GET','POST'])
@is_logged_in
def importNewForm():
    #crear un nuevo formulario a partir de un archivo de excel
    response={}
    try:
        data=request.form.to_dict()
        success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
        if success:
            if allowed:
                files=request.files
                file_path=cfg.uploaded_forms_files_path
                file=files[data['file_name']]
                filename = secure_filename(file.filename)
                file.save(os.path.join(file_path, filename))
                read_file=load_workbook(os.path.join(file_path,filename))
                ws = read_file.worksheets[0]
                # ws=read_file['Sheet1']

                form={
                    'project_id':data['project_id'],
                    'name':data['name'],
                    'columns_number':int(ws.max_column),
                    'rows':int(ws.max_row),
                    'created_by':int(data['user_id']),
                    'create_date':'now',
                    'folder_id':int(data['folder_id']),
                    'status_id':2,
                    'user_last_update':int(data['user_id']),
                    'last_updated':'now'
                }

                app.logger.info(ws.max_column)
                columns=[]
                for x in range(1,int(ws.max_column)+1):
                    column={
                        'order':x,
                        'name':ws.cell(row=1, column=x).value
                    }
                    if ws.cell(row=2, column=x).value is None:
                        column['editable']=True
                    else:
                        column['editable']=False
                    columns.append(column)
                form['columns']=str(columns)
                db.insert('project.form',form)

                table_name='project_%s_form_%s'%(form['project_id'],form['form_id'])


                columns_str=""
                for c in columns:
                    columns_str+=" ,col_%s text default ''"%c['order']
                db.query("""
                    CREATE TABLE form.%s(
                        entry_id serial not null primary key,
                        form_id integer not null,
                        project_id integer not null
                        %s
                    );
                """%(table_name,columns_str))

                for i in range(0,int(form['rows'])-1):
                    ins={
                        'form_id':form['form_id'],
                        'project_id':form['project_id']
                    }
                    for c in columns:
                        if c['editable']==False:
                            if ws.cell(row=i+2, column=c['order']).value is not None:
                                if type(ws.cell(row=i+2, column=c['order']).value)=='unicode':
                                    ins['col_%s'%c['order']]=(ws.cell(row=i+2, column=c['order']).value).encode('utf-8')
                                else:
                                    ins['col_%s'%c['order']]=ws.cell(row=i+2, column=c['order']).value
                            else:
                                ins['col_%s'%c['order']]=''
                    db.insert('form.%s'%table_name,ins)
                response['success']=True
                response['msg_response']='El formulario ha sido agregado.'
            else:
                response['success']=False
                response['msg_response']='No tienes permisos para realizar esta acción.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar validar la información.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/checkAllowedDownload', methods=['GET','POST'])
@is_logged_in
def checkAllowedDownload():
    #revisa si tiene permisos para descargar el formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'download_forms'})
                if success:
                    if allowed:
                        status=db.query("""
                            select status_id from project.form where form_id=%s
                        """%data['form_id']).dictresult()
                        #debe ser el último status
                        # if int(status[0]['status_id'])==7:
                        if int(status[0]['status_id'])>2:
                            manager=db.query("""
                                select * from project.project where project_id=%s and (manager=%s or partner=%s)
                            """%(data['project_id'],data['user_id'],data['user_id'])).dictresult()
                            if manager!=[]:
                                response['allowed']=True
                            else:
                                users=db.query("""
                                    select a.* from project.project_users a, system.user b where
                                    b.download_forms=True and a.user_id=b.user_id and a.project_id=%s and a.user_id=%s
                                """%(data['project_id'],data['user_id'])).dictresult()
                                if users!=[]:
                                    response['allowed']=True
                                else:
                                    response['allowed']=False
                                    response['msg_response']='No tienes permisos para descargar este formulario.'
                        else:
                            response['allowed']=False
                            response['msg_response']='Este formulario no puede ser descargado aún.'
                        response['success']=True
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/doDownloadResolvedForm', methods=['GET','POST'])
@is_logged_in
def downloadResolvedForm():
    #crear archivo de excel de formulario resuelto para su descarga
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'download_forms'})
                if success:
                    if allowed:
                        project=db.query("""
                            select
                                name,
                                company_name,
                                (select a.name from system.user a where a.user_id=manager) as manager,
                                (select a.name from system.user a where a.user_id=partner) as partner,
                                comments
                            from
                                project.project
                            where project_id=%s
                        """%data['project_id']).dictresult()[0]
                        form=db.query("""
                            select
                                name,
                                columns_number,
                                rows,
                                status_id,
                                columns,
                                (select a.name from system.user a where a.user_id=assigned_to) as assigned_to,
                                to_char(resolved_date,'DD/MM/YYYY HH24:MI:SS') as resolved_date,
                                to_char(first_revision_date,'DD/MM/YYYY HH24:MI:SS') as first_revision_date
                                --,revisions
                            from
                                project.form
                            where
                                form_id=%s
                        """%data['form_id']).dictresult()[0]
                        wb = Workbook()
                        ws = wb.create_sheet('Hoja1',0)

                        company_style=Font(
                            name='Times New Roman',
                            size=14,
                            bold=True,
                            italic=False,
                            color='FF000080'
                        )
                        project_style=Font(
                            name='Times New Roman',
                            size=14,
                            bold=False,
                            italic=True,
                            color='FF000080'
                        )
                        formname_style=Font(
                            name='Times New Roman',
                            size=14,
                            bold=False,
                            italic=False,
                            color='FF000080'
                        )

                        header_style=NamedStyle(name='header_style')
                        header_style.font=Font(
                            name='Times New Roman',
                            size=14,
                            bold=True,
                            italic=False,
                            color='FFFFFFFF'
                        )
                        header_style.fill=PatternFill("solid", fgColor="FF000099")
                        header_style.alignment=Alignment(horizontal='center')
                        header_style.border=Border(
                            left=Side(border_style='thin',color='FF000000'),
                            right=Side(border_style='thin',color='FF000000'),
                            top=Side(border_style='thin',color='FF000000'),
                            bottom=Side(border_style='thin',color='FF000000')
                        )

                        content_style=NamedStyle(name='content_style')
                        content_style.font=Font(
                            name='Times New Roman',
                            size=12,
                            bold=False,
                            italic=False,
                            color='FF000000'
                        )
                        content_style.alignment=Alignment(
                            horizontal='justify',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)
                        # wrap_text=True,shrink_to_fit=False,vertical='justify',horizontal='left')
                        content_style.border=Border(
                            left=Side(border_style='thin',color='FF000000'),
                            right=Side(border_style='thin',color='FF000000'),
                            top=Side(border_style='thin',color='FF000000'),
                            bottom=Side(border_style='thin',color='FF000000')
                        )

                        ws.sheet_view.showGridLines = False #ocultar líneas

                        ws['B2']=project['company_name']
                        ws['B2'].font=company_style
                        ws['B3']=project['name']
                        ws['B3'].font=project_style
                        ws['B4']=form['name']
                        ws['B4'].font=formname_style

                        columns=eval(form['columns'])
                        col_num=2
                        for c in columns:
                            ws.cell(column=col_num,row=6,value=c['name'])
                            ws.cell(column=col_num,row=6).style=header_style
                            col_num+=1
                        ws.cell(column=col_num,row=6,value='Revisión')
                        ws.cell(column=col_num,row=6).style=header_style

                        form_info=db.query("""
                            select * from form.project_%s_form_%s order by entry_id
                        """%(data['project_id'],data['form_id'])).dictresult()
                        row=7

                        for x in form_info:
                            keys=sorted(x.iteritems())
                            col_num=2
                            for k in keys:
                                if k[0].split('_')[0]=='col':
                                    ws.cell(column=col_num,row=row,value=k[1].decode('utf-8'))
                                    ws.cell(column=col_num,row=row).style=content_style

                                    col_num+=1
                            ws.cell(column=col_num,row=row,value=x['rev_1'].decode('utf-8'))
                            ws.cell(column=col_num,row=row).style=content_style
                            row+=1

                        last_info_row=row

                        #comprueba si el status es cerrado para agregar datos de quien realizó y revisó el formulario
                        if int(form['status_id'])==7:

                            #agregar quién realizó el formulario
                            revisions=db.query("""
                                select (select a.name from system.user a where a.user_id=b.user_id) as user_name,
                                b.revision_number, to_char(b.revision_date,'DD/MM/YYYY HH24:MI:SS') as revision_date
                                from project.form_revisions b where b.form_id=%s order by b.revision_number asc
                            """%data['form_id']).dictresult()
                            bd = Side(style='thick', color="000000")
                            th = Side(style='thin', color="000000")

                            col_num+=2
                            ws.cell(column=col_num,row=6).border=Border(left=bd,top=bd,bottom=th)
                            ws.cell(column=col_num+1,row=6,value='Nombre')
                            ws.cell(column=col_num+1,row=6).font=Font(name='Times New Roman', size=12, bold=True)
                            ws.cell(column=col_num+1,row=6).border=Border(top=bd,bottom=th)
                            ws.cell(column=col_num+2,row=6,value='Fecha')
                            ws.cell(column=col_num+2,row=6).font=Font(name='Times New Roman', size=12, bold=True)
                            ws.cell(column=col_num+2,row=6).border=Border(top=bd,right=bd,bottom=th)

                            ws.cell(column=col_num,row=7,value='Encargado')
                            ws.cell(column=col_num,row=7).font=Font(name='Times New Roman', size=12, bold=True)
                            ws.cell(column=col_num,row=7).border=Border(left=bd)
                            ws.cell(column=col_num+1,row=7,value=form['assigned_to'])
                            ws.cell(column=col_num+1,row=7).font=Font(name='Times New Roman', size=12, bold=False)
                            ws.cell(column=col_num+2,row=7,value=form['resolved_date'])
                            ws.cell(column=col_num+2,row=7).font=Font(name='Times New Roman', size=12, bold=False)
                            ws.cell(column=col_num+2,row=7).border=Border(right=bd)

                            current_row=8
                            for r in revisions:
                                ws.cell(column=col_num,row=current_row,value='Revisor')
                                ws.cell(column=col_num,row=current_row).font=Font(name='Times New Roman', size=12, bold=True)
                                ws.cell(column=col_num,row=current_row).border=Border(left=bd)
                                # revisor=db.query("""
                                #     select name from system.user where user_id=%s
                                # """%int(form['revisions'].split(",")[0].split(":")[1])).dictresult()[0]
                                ws.cell(column=col_num+1,row=current_row,value=r['user_name'])
                                ws.cell(column=col_num+1,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
                                ws.cell(column=col_num+2,row=current_row,value=r['revision_date'])
                                ws.cell(column=col_num+2,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
                                ws.cell(column=col_num+2,row=current_row).border=Border(right=bd)
                                current_row+=1

                            ws.cell(column=col_num,row=current_row,value='Gerente')
                            ws.cell(column=col_num,row=current_row).font=Font(name='Times New Roman', size=12, bold=True)
                            ws.cell(column=col_num,row=current_row).border=Border(left=bd)
                            ws.cell(column=col_num+1,row=current_row,value=project['manager'])
                            ws.cell(column=col_num+1,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
                            ws.cell(column=col_num+2,row=current_row).border=Border(right=bd)
                            current_row+=1

                            ws.cell(column=col_num,row=current_row,value='Socio')
                            ws.cell(column=col_num,row=current_row).font=Font(name='Times New Roman', size=12, bold=True)
                            ws.cell(column=col_num,row=current_row).border=Border(left=bd,bottom=bd)
                            ws.cell(column=col_num+1,row=current_row).border=Border(bottom=bd)
                            ws.cell(column=col_num+1,row=current_row,value=project['partner'])
                            ws.cell(column=col_num+1,row=current_row).font=Font(name='Times New Roman', size=12, bold=False)
                            ws.cell(column=col_num+2,row=current_row).border=Border(right=bd,bottom=bd)


                            for column_cells in ws.columns:
                                length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                                ws.column_dimensions[column_cells[0].column].width = length


                            #revisar si hay observaciones del formulario
                            comments=db.query("""
                                select b.comment, to_char(b.created,'DD/MM/YYYY HH24:MI:SS') as created, (select a.name from system.user a where a.user_id=b.user_id) as user_name from project.form_comments b where b.form_id=%s
                            """%data['form_id']).dictresult()
                            if comments!=[]:
                                comm_header_style=NamedStyle(name='comm_header_style')
                                comm_header_style.font=Font(
                                    name='Times New Roman',
                                    size=14,
                                    bold=True,
                                    italic=False,
                                    color='FFFFFFFF'
                                )
                                comm_header_style.fill=PatternFill("solid", fgColor="FF7082D4")
                                comm_header_style.alignment=Alignment(horizontal='center')
                                comm_header_style.border=Border(
                                    left=Side(border_style='thin',color='FF000000'),
                                    right=Side(border_style='thin',color='FF000000'),
                                    top=Side(border_style='thin',color='FF000000'),
                                    bottom=Side(border_style='thin',color='FF000000')
                                )

                                wo = wb.create_sheet('Observaciones',1)
                                row=2
                                wo.cell(column=2, row=row, value='Por')
                                wo.cell(column=2, row=row).style=comm_header_style
                                wo.cell(column=3, row=row, value='Fecha')
                                wo.cell(column=3, row=row).style=comm_header_style
                                wo.cell(column=4, row=row, value='Observación')
                                wo.cell(column=4, row=row).style=comm_header_style
                                row+=1
                                for x in comments:
                                    wo.cell(column=2,row=row,value=x['user_name'])
                                    wo.cell(column=2,row=row).style=content_style
                                    wo.cell(column=3,row=row,value=x['created'])
                                    wo.cell(column=3,row=row).style=content_style
                                    wo.cell(column=4,row=row,value=x['comment'])
                                    wo.cell(column=4,row=row).style=content_style
                                    row+=1

                                for column_cells in wo.columns:
                                    length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                                    wo.column_dimensions[column_cells[0].column].width = length





                        #Descarga de archivo
                        time=strftime("%H_%M_%S", gmtime())
                        path=os.path.join(cfg.uploaded_forms_files_path,'Reporte%s.xlsx'%time)
                        wb.save(path)
                        response['success']=True
                        response['msg_response']='El formulario ha sido generado.'
                        response['filename']='/project/downloadFile/report/Reporte%s.xlsx'%time
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/downloadFile/<type>/<filename>', methods=['GET','POST'])
@is_logged_in
def downloadReport(type,filename):
    response={}
    try:
        if type=='report':
            path=os.path.join(cfg.uploaded_forms_files_path,filename)
        elif type=='form_zip':
            path=os.path.join(cfg.download_zip_path,filename)
        name="%s"%filename
        return send_file(path,attachment_filename=name)
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
        return response

@bp.route('/uploadFormZipFile', methods=['GET','POST'])
@is_logged_in
def uploadFormZipFile():
    #cargar archivo zip a formulario
    response={}
    try:
        if request.method=='POST':
            data=request.form.to_dict()
            success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
            if success:
                if allowed:
                    file=request.files[data['file_name'].encode('utf-8')]

                    filename=secure_filename(file.filename)
                    folder_path=os.path.join(cfg.zip_main_folder,'project_%s'%int(data['project_id']),'form_%s'%int(data['form_id']))
                    if not os.path.exists(folder_path):
                        os.makedirs(folder_path)
                    # else:
                    #     existing_files = glob.glob(folder_path+'/*')
                    #     for f in existing_files:
                    #         os.remove(f)

                    #revisar si ya existe un archivo con ese nombre
                    file_exists=db.query("""
                        select count(*) from project.form_files where file_name='%s' and form_id=%s and project_id=%s
                    """%(filename,data['form_id'],data['project_id'])).dictresult()

                    if file_exists[0]['count']==0:
                        file.save(os.path.join(folder_path,filename))
                        file_insert={
                            'file_name':filename,
                            'form_id':data['form_id'],
                            'project_id':data['project_id'],
                            'uploaded_by':data['user_id'],
                            'upload_date':'now',
                            'file_name_display':data['file_name'].encode('utf-8')
                        }
                        db.insert('project.form_files',file_insert)
                        # db.query("""
                        #     update project.form
                        #     set zip_file_name='%s',
                        #     zip_last_upload_user=%s,
                        #     zip_last_upload_date='now'
                        #     where form_id=%s and project_id=%s
                        # """%(filename,data['user_id'],data['form_id'],data['project_id']))
                        response['success']=True
                        response['msg_response']='El archivo ha sido cargado.'
                    else:
                        response['success']=False
                        response['msg_response']='Ya existe un archivo con ese nombre, favor de cambiarlo.'
                else:
                    response['success']=False
                    response['msg_response']='No tienes permisos para realizar esta acción.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/allowedFormZip',methods=['GET','POST'])
@is_logged_in
def allowedFormZip():
    #verifica si tiene permisos para subir o descargar el archivo zip del formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        response['success']=True
                        form_users=db.query("""
                            select assigned_to, status_id,
                            --revisions,
                            project_id from project.form
                            where form_id=%s
                        """%data['form_id']).dictresult()[0]
                        if (form_users['status_id']<7 and data['from']=='upload') or data['from']=='download':
                            if int(data['user_id'])==int(form_users['assigned_to']):
                                response['allowed']=True
                            else:
                                revisions=db.query("""
                                    select user_id from project.form_revisions where form_id=%s
                                """%data['form_id']).dictresult()
                                # rev_list=form_users['revisions'].split(",")
                                # for x in rev_list:
                                for r in revisions:
                                    # if int(x.split(":")[1])==int(data['user_id']):
                                    if int(r['user_id'])==int(data['user_id']):
                                        response['allowed']=True
                                        break
                                    else:
                                        response['allowed']=False
                                        if data['from']=='upload':
                                            response['msg_response']='No tienes permitido subir un archivo relacionado a este formulario.'
                                        else:
                                            response['msg_response']='No tienes permitido descargar archivos de este formulario.'
                                if response['allowed']==False:
                                    project_users=db.query("""
                                        select manager,partner from project.project
                                        where project_id=%s
                                    """%form_users['project_id']).dictresult()[0]
                                    if int(data['user_id'])==int(project_users['manager']) or int(data['user_id'])==int(project_users['partner']):
                                        response['allowed']=True
                                    else:
                                        workspace_id=db.query("""
                                            select workspace_id from system.user where user_id=%s
                                        """%project_users['manager']).dictresult()[0]
                                        consultants=db.query("""
                                            select consultants from system.workspace where workspace_id=%s
                                        """%workspace_id['workspace_id']).dictresult()[0]
                                        cons=consultants['consultants'].split(",")
                                        if str(data['user_id']) in cons:
                                            response['allowed']=True
                                        else:
                                            response['allowed']=False
                                            if data['from']=='upload':
                                                response['msg_response']='No tienes permitido subir un archivo relacionado a este formulario.'
                                            else:
                                                response['msg_response']='No tienes permitido descargar archivos de este formulario.'
                        else:
                            response['allowed']=False
                            response['msg_response']='Este formulario se encuentra cerrado, por lo tanto no puede ser cargado ningún archivo nuevo.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getZipDownloadLink', methods=["GET","POST"])
@is_logged_in
def getZipDownloadLink():
    #obtiene el link para la descarga del archivo zip del formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        zip_file=db.query("""
                            select zip_file_name
                            from project.form
                            where form_id=%s
                        """%data['form_id']).dictresult()[0]
                        response['success']=True
                        if zip_file['zip_file_name']!='':
                            file_link='/project/downloadZipFile/%s/%s/%s'%(data['project_id'],data['form_id'],zip_file['zip_file_name'])
                            response['file']=file_link
                            response['has_file']=True
                            response['msg_response']='El archivo se encuentra disponible.'
                        else:
                            response['has_file']=False
                            response['msg_response']='No hay ningún archivo por descargar.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/downloadZipFile/<project_id>/<form_id>/<filename>', methods=['GET','POST'])
@is_logged_in
def downloadZipFile(project_id,form_id,filename):
    #realiza descarga de archivo zip
    #no requiere validación de permisos
    response={}
    try:
        path=os.path.join(cfg.zip_main_folder,'project_%s'%project_id,'form_%s'%form_id,filename)
        name="%s"%filename
        return send_file(path,attachment_filename=name)
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
        return response

@bp.route('/getUsersForRevision', methods=['GET','POST'])
@is_logged_in
def getUsersForRevision():
    #obtiene los usuarios a quienes se les puede regresar el formulario o pasar para revisión
    #no requiere validación de permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                revisions=db.query("""
                    select a.user_id, b.name, a.revision_number, a.currently_assigned
                    from system.user b,
                    project.form_revisions a
                    where form_id=%s
                    and a.user_id=b.user_id order by revision_number asc
                """%data['form_id']).dictresult()
                assignee=db.query("""
                    select a.assigned_to as user_id, b.name from system.user b, project.form a
                    where a.form_id=%s
                    and a.assigned_to=b.user_id
                """%data['form_id']).dictresult()[0]
                send_to=[]
                return_to=[]
                assignee['name']='%s - Asignado'%assignee['name']
                return_to.append(assignee)
                one_more=False
                for r in revisions:
                    if one_more==True:
                        send_to.append({'user_id':r['user_id'],'name':"%s - Revisión %s"%(r['name'],r['revision_number'])})
                        break
                    else:
                        if r['currently_assigned']==True:
                            one_more=True
                        else:
                            return_to.append({'user_id':r['user_id'],'name':"%s - Revisión %s"%(r['name'],r['revision_number'])})
                if send_to==[]:
                    send_to.append({'user_id':-101,'name':'Cerrar formulario'})

                response['success']=True
                response['send_to']=send_to
                response['return_to']=return_to
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/doRevision', methods=['GET','POST'])
@is_logged_in
def doRevision():
    #realizar revisión de formulario, ya sea regresarlo, mandarlo a la siguiente revisión o cerrarlo
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                app.logger.info(data)
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'resolve_forms'})
                if success:
                    if allowed:
                        if int(data['user_to'])==-101: #revisar si va a cerrar el formulario
                            #buscar el usuario que tiene la última revisión del formulario
                            last_revision=db.query("""
                                select revision_id, user_id
                                from project.form_revisions
                                where form_id=%s order by revision_number desc limit 1
                            """%data['form_id']).dictresult()[0]
                            if int(last_revision['user_id'])==int(data['user_id']): # si es el usuario de la última revisión, se cambia el status del formulario a cerrado
                                db.query("""
                                    update project.form set
                                    status_id=7,
                                    user_last_update=%s,
                                    last_updated='now'
                                    where form_id=%s
                                """%(data['user_id'],data['form_id']))
                                #se actualiza la fecha de la revisión
                                db.query("""
                                    update project.form_revisions
                                    set revision_date='now'
                                    where form_id=%s and user_id=%s
                                """%(data['form_id'],data['user_id']))

                                users_notif=db.query("""
                                    select user_id from project.form_revisions where user_id!=%s and form_id=%s
                                    union select assigned_to as user_id from project.form where form_id=%s
                                """%(data['user_id'],data['form_id'],data['form_id'])).dictresult()
                                info={
                                    'form_id':data['form_id'],
                                    'project_id':data['project_id'],
                                    'msg':data['msg']
                                }
                                for un in users_notif:
                                    info['user_to']=un['user_id']
                                    # se envía notificación a usuarios correspondientes
                                    GF.createNotification('close_form',info)
                                send_mail=db.query("""
                                    select notify_resolved from project.form where form_id=%s
                                """%data['form_id']).dictresult()[0]
                                if send_mail['notify_resolved']==True:
                                    GF.sendMailNotification({'form_id':data['form_id'],'type':'resolved_form'})
                        else:
                            #si no se va a cerrar el formulario
                            #se revisa si se va a regresar al usuario asignado a resolver el formulario
                            check_assignee=db.query("""
                                select assigned_to from project.form
                                where form_id=%s
                            """%data['form_id']).dictresult()[0]
                            if int(check_assignee['assigned_to'])==int(data['user_to']):
                                #si es el usuario asignado, se cambia status a publicado
                                db.query("""
                                    update project.form
                                    set status_id=3, user_last_update=%s,
                                    last_updated='now' where form_id=%s
                                """%(data['user_id'],data['form_id']))
                                #actualizar fecha en que se regresó la revisión en registro de usuario que regresó el formulario
                                db.query("""
                                    update project.form_revisions set returned_revision_date='now' where form_id=%s
                                    and currently_assigned=True
                                """%data['form_id'])
                                #se actualiza a quien está asignada la revisión
                                db.query("""
                                    update project.form_revisions set
                                    currently_assigned=False, revision_date='1900-01-01 00:00:00' where form_id=%s
                                """%(data['form_id']))

                                notif_info={'project_id':data['project_id'],'form_id':data['form_id'],'msg':data['msg']}
                                GF.createNotification('return_form',notif_info)
                                assignee_info=db.query("""
                                    select a.name, a.email, b.name as form_name from system.user a, project.form b
                                    where a.user_id=b.assigned_to and b.form_id=%s
                                """%data['form_id']).dictresult()[0]
                                revisor_info=db.query("""
                                    select name, email from system.user where user_id=%s
                                """%data['user_id']).dictresult()
                                assignee_info['revisor_name']=revisor_info[0]['name']
                                assignee_info['comments']=data['msg'].encode('utf-8')
                                assignee_info['top_img']=cfg.img_return_form
                                assignee_info['bottom_img']=cfg.img_rb_logo
                                link=os.path.join(cfg.host,'project',str(cfg.project_factor*int(data['project_id'])),str(data['form_id']))
                                assignee_info['link']='<a href="%s"> aqu&iacute;</a>'%link
                                template='<p style="text-align: center;"><img src="{top_img}" alt="" width="50" height="50"></p><p><span style="font-family: Verdana, Geneva, sans-serif; color: rgb(93, 93, 92);">Estimado usuario:</span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Le informamos que el usuario {revisor_name} ha revisado el programa de trabajo {form_name} y ha decidido asignarlo a usted nuevamente.&nbsp;</span></span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Ha incluido los siguientes comentarios:</span></span></p><p style="margin-left: 20px;"><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;"><em>{comments}</em></span></span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Para acceder al programa de trabajo, dar click {link}.</span></span></p><p><br></p><p><img src="{bottom_img}" alt="" width="250" height="70"></p>'
                                body=template.format(**assignee_info)
                                GF.sendMail('Revisión completada (regresado)',body,assignee_info['email'])
                            else:
                                #en caso de que se cambie de revisor
                                current_revision=db.query("""
                                    select * from project.form_revisions where
                                    form_id=%s and currently_assigned=True
                                """%data['form_id']).dictresult()[0]
                                app.logger.info("current_revision")
                                app.logger.info(current_revision)
                                #se actualiza quien y cuando fue la última vez que se actualizó el formulario
                                db.query("""
                                    update project.form set
                                    user_last_update=%s, last_updated='now', status_id=6
                                    where form_id=%s
                                """%(data['user_id'],data['form_id']))
                                user_update=""
                                if int(current_revision['user_id'])!=int(data['user_id']):
                                    user_update=",user_update=%s"%data['user_id']
                                #se actualiza la fecha en que se realizó la revisión, y se cambia currently_assigned a false
                                db.query("""
                                    update project.form_revisions set currently_assigned=False,
                                    revision_date='now' %s where form_id=%s and currently_assigned=True
                                """%(user_update,data['form_id']))
                                #se asigna el nuevo revisor
                                db.query("""
                                    update project.form_revisions set currently_assigned=True
                                    where form_id=%s and user_id=%s
                                """%(data['form_id'],data['user_to']))
                                #cuando se regresa a un revisor anterior
                                if data['to']=='return':
                                    new_assigned=db.query("""
                                        select revision_number from project.form_revisions where user_id=%s and form_id=%s
                                    """%(data['user_to'],data['form_id'])).dictresult()[0]

                                    db.query("""
                                        update project.form_revisions set revision_date='1900-01-01 00:00:00' where
                                        form_id=%s and revision_number>=%s
                                    """%(data['form_id'],new_assigned['revision_number']))
                                    db.query("""
                                        update project.form_revisions set returned_revision_date='now' where revision_id=%s
                                    """%current_revision['revision_id'])
                                notif_info={'project_id':data['project_id'],'form_id':data['form_id'],'msg':data['msg'],'user_to':data['user_to']}
                                GF.createNotification('send_next_revision',notif_info)

                                form_info=db.query("""
                                    select name as form_name from project.form where form_id=%s
                                """%data['form_id']).dictresult()[0]
                                form_info['revisor_name']=db.query("""
                                    select name from system.user where user_id=%s
                                """%data['user_id']).dictresult()[0]['name']
                                form_info['email']=db.query("""
                                    select email from system.user where user_id=%s
                                """%data['user_to']).dictresult()[0]['email']
                                form_info['comments']=data['msg'].encode('utf-8')
                                form_info['top_img']=cfg.img_revision_form
                                form_info['bottom_img']=cfg.img_rb_logo
                                link=os.path.join(cfg.host,'project',str(cfg.project_factor*int(data['project_id'])),str(data['form_id']))
                                form_info['link']='<a href="%s"> aqu&iacute;</a>'%link
                                template='<p style="text-align: center;"><img src="{top_img}" alt="" width="50" height="50"></p><p><span style="font-family: Verdana, Geneva, sans-serif; color: rgb(93, 93, 92);">Estimado usuario:</span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Le informamos que el usuario {revisor_name} ha revisado el programa de trabajo {form_name} y ha decidido asignarlo a usted para su revisi&oacute;n.&nbsp;</span></span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Ha incluido los siguientes comentarios:</span></span></p><p style="margin-left: 20px;"><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;"><em>{comments}</em></span></span></p><p><span style="color: rgb(93, 93, 92);"><span style="font-family: Verdana, Geneva, sans-serif;">Para acceder al programa de trabajo, dar click {link}.</span></span></p><p><img src="{bottom_img}" alt="" width="250" height="70"></p><p><br></p>'
                                body=template.format(**form_info)
                                GF.sendMail('Enviado a revisión',body,form_info['email'])


                        response['success']=True
                        response['msg_response']='El formulario ha sido actualizado.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getPrintingInfo', methods=['GET','POST'])
@is_logged_in
def getPrintingInfo():
    #regresa la información del formulario para su vista de impresión
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'download_forms'})
                if success:
                    if allowed:
                        form=db.query("""
                            select name, (select a.name from system.user a where a.user_id=assigned_to) as assigned_to,
                            to_char(resolved_date,'DD/MM/YYYY HH24:MI:SS') as resolved_date
                            from project.form where form_id=%s
                        """%data['form_id']).dictresult()[0]
                        project=db.query("""
                            select name, company_name, (select a.name from system.user a where a.user_id=manager)  as manager,
                            (select a.name from system.user a where a.user_id=partner) as partner
                            from project.project where project_id=%s
                        """%data['project_id']).dictresult()[0]
                        revisions=db.query("""
                            select user_id, revision_number, to_char(revision_date,'DD/MM/YYYY HH24:MI:SS') as revision_date
                            from project.form_revisions where form_id=%s order by revision_number asc
                        """%data['form_id']).dictresult()
                        html='<p style="font-size:1.6em;">{name}</p><p style="font-size:1.4em;">{company_name}</p>'.format(**project)
                        html+='<p style="font-size:1.2em; font-weight:bold;">{name}</p><p><b>Realizado por:</b> {assigned_to} - {resolved_date}</p>'.format(**form)
                        for x in revisions:
                            name=db.query("select name from system.user where user_id=%s"%x['user_id']).dictresult()[0]['name']
                            x['name']=name
                            html+='<p><b>Revisión {revision_number}:</b> {name} - {revision_date}</p>'.format(**x)
                        html+='<p><b>Gerente:</b> {manager}</p><p><b>Socio:</b> {partner}</p>'.format(**project)
                        response['success']=True
                        response['html']=html
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/checkRightPanelPermission', methods=['GET','POST'])
@is_logged_in
def checkRightPanelPermission():
    #revisar si tiene permisos para crear formularios, para mostrar el panel derecho de formularios por revisar y formularios por publicar
    #no requiere validación
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                permission=db.query("""
                    select create_forms from system.user where user_id=%s
                """%data['user_id']).dictresult()[0]
                if permission['create_forms']==True:
                    response['allowed']=True
                else:
                    response['allowed']=False
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/cloneForm', methods=['GET','POST'])
@is_logged_in
def cloneForm():
    #clonar un formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        ofi=db.query("""
                            select * from project.form where form_id=%s
                        """%data['old_form_id']).dictresult()[0]
                        new_form={
                            'project_id':ofi['project_id'],
                            'name':data['form_name'],
                            'columns_number':ofi['columns_number'],
                            'rows':ofi['rows'],
                            'columns':ofi['columns'],
                            'created_by':data['user_id'],
                            'create_date':'now',
                            'folder_id':data['new_folder_id'],
                            'status_id':2,
                            'user_last_update':data['user_id'],
                            'last_updated':'now'
                        }
                        inserted_form=db.insert('project.form',new_form)

                        columns=eval(ofi['columns'])

                        table_name='project_%s_form_%s'%(inserted_form['project_id'],inserted_form['form_id'])
                        columns_str=""
                        including_columns=[]
                        for x in columns:
                            columns_str+=" ,col_%s text default ''"%x['order']
                            if x['editable']==False:
                                including_columns.append("col_%s"%x['order'])

                        db.query("""
                            CREATE TABLE form.%s(
                                entry_id serial not null primary key,
                                form_id integer not null,
                                project_id integer not null
                                %s
                            );
                        """%(table_name,columns_str))

                        old_form_table='project_%s_form_%s'%(ofi['project_id'],ofi['form_id'])
                        old_info=db.query("""
                            select * from form.%s order by entry_id
                        """%old_form_table).dictresult()
                        for oi in old_info:
                            entry={
                                'form_id':inserted_form['form_id'],
                                'project_id':inserted_form['project_id']
                            }
                            if including_columns!=[]:
                                for ic in including_columns:
                                    entry[ic]=oi[ic]
                            db.insert('form.%s'%table_name,entry)
                        response['success']=True
                        response['msg_response']='El formulario ha sido clonado exitosamente.'
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getFormDocuments', methods=['GET','POST'])
@is_logged_in
def getFormDocuments():
    #obtiene los documentos que han sido cargados a un formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                response['success']=True
                documents=db.query("""
                    select file_id, file_name, file_name_display from project.form_files
                    where form_id=%s and project_id=%s
                    order by upload_date desc
                """%(data['form_id'],data['project_id'])).dictresult()
                html=''
                extensions={'.xlsx':'icon-excel','.pptx':'icon-powerpoint','.pdf':'icon-pdf','.docx':'icon-word','.zip':'icon-zip'}
                if documents!=[]:
                    for x in documents:
                        name,ext=os.path.splitext(x['file_name'])
                        if ext in extensions:
                            doc_class=extensions[ext]
                        else:
                            doc_class='icon-generic-file'
                        file_display,ext=os.path.splitext(x['file_name_display'])
                        file_link='/project/downloadZipFile/%s/%s/%s'%(data['project_id'],data['form_id'],x['file_name'])
                        html+='<div class="download-icon-div"><input type="checkbox" class="checkbox-download-files" data-document="%s"><div style="display:grid;"><a href="%s" target="_blank" data-toggle="tooltip" title="%s" class="download-file"><i class="%s"></i></a><span class="spn-icon-text" data-toggle="tooltip" title="%s">%s</span></div></div>'%(x['file_id'],file_link,file_display.decode('utf8'),doc_class,file_display.decode('utf8'),file_display.decode('utf8'))
                        response['data']=html
                else:
                    response['data']='<h4 style="color:#737a80;">No hay documentos cargados<h4>'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getDownloadFolderLink', methods=['GET','POST'])
@is_logged_in
def getDownloadFolderLink():
    #crear carpeta zip con archivos a descargar y generar el link de descarga
    #no requiere validación porque para abrir la ventana de descarga, se validan los permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:

                file_ids=','.join(str(e) for e in data['file_list'])
                files=db.query("""
                    select file_name from project.form_files
                    where file_id in (%s) and project_id=%s and form_id=%s
                """%(file_ids,data['project_id'],data['form_id'])).dictresult()

                temp_folder='Descarga_%s.zip'%strftime("%H_%M_%S", localtime())
                folder_path=os.path.join(cfg.download_zip_path,temp_folder)
                origin_path=os.path.join(cfg.zip_main_folder,'project_%s'%int(data['project_id']),'form_%s'%int(data['form_id']))
                with zipfile.ZipFile(folder_path, 'w') as new_zip:
                    for x in files:
                        new_zip.write(os.path.join(origin_path,x['file_name']),x['file_name'])
                    new_zip.close()
                response['success']=True
                response['link']=temp_folder
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/deleteFormFile', methods=['GET','POST'])
@is_logged_in
def deleteFormFile():
    #eliminar archivo(s) de un formulario
    #no requiere validación, porque si puede acceder a los archivos descargados, tiene permisos para eliminar el o los archivos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                form_status=db.query("""
                    select status_id from project.form where form_id=%s
                """%data['form_id']).dictresult()[0]
                if int(form_status['status_id'])!=7:
                    file_ids=','.join(str(e) for e in data['files'])
                    files=db.query("""
                        select file_name from project.form_files
                        where file_id in (%s) and form_id=%s and project_id=%s
                    """%(file_ids,data['form_id'],data['project_id'])).dictresult()
                    for f in files:
                        os.remove(os.path.join(cfg.zip_main_folder,'project_%s'%data['project_id'],'form_%s'%data['form_id'],f['file_name']))
                    db.query("""
                        delete from project.form_files
                        where file_id in (%s)
                    """%file_ids)
                    response['success']=True
                    if len(data['files'])==1:
                        response['msg_response']='El archivo ha sido eliminado.'
                    else:
                        response['msg_response']='Los archivos han sido eliminados.'
                else:
                    response['success']=False
                    response['msg_response']='No es posible eliminar archivos de un formulario cerrado.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/deleteMenuElements', methods=['GET','POST'])
@is_logged_in
def deleteMenuElements():
    #eliminar carpetas o formularios listados en el proyecto
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'delete_foldersandforms'})
                if success:
                    if allowed:
                        folder_list=[]
                        form_list=[]
                        if data['folders']!=[]:
                            for y in data['folders']:
                                folder_list,form_list=getFormsToDelete(y,folder_list,form_list)
                        if data['forms']!=[]:
                            for f in data['forms']:
                                form_list.append(f)

                        if form_list!=[]:
                            for x in form_list:
                                #1. eliminar comentarios
                                db.query("""
                                    delete from project.form_comments
                                    where form_id=%s
                                """%(x))
                                #2. eliminar archivos (tabla y en carpeta)
                                db.query("""
                                    delete from project.form_files
                                    where form_id=%s
                                """%x)
                                if os.path.exists(os.path.join(cfg.zip_main_folder,'project_%s'%data['project_id'],'form_%s'%x)):
                                    shutil.rmtree(os.path.join(cfg.zip_main_folder,'project_%s'%data['project_id'],'form_%s'%x))
                                #3. notificaciones
                                db.query("""
                                    delete from project.notification
                                    where form_id=%s and project_id=%s
                                """%(x,data['project_id']))
                                #4. revisores
                                db.query("""
                                    delete from project.form_revisions
                                    where form_id=%s
                                """%x)
                                #5. tabla form
                                db.query("""
                                    drop table form.project_%s_form_%s ;
                                """%(data['project_id'],x))
                                #6. registro de formulario de project.form
                                db.query("""
                                    delete from project.form where form_id=%s
                                """%x)

                        if folder_list!=[]:
                            folder_str=','.join(str(e) for e in folder_list)
                            db.query("""
                                delete from project.folder where project_id=%s and folder_id in (%s)
                            """%(data['project_id'],folder_str))

                        response['success']=True

                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

def getFormsToDelete(folder_id,folder_list,form_list):
    response={}
    #1. obtener subfolders con folder_id como parent_id
    #2. al final buscar todos los formularios de las carpetas del listado de carpetas
    try:
        if folder_id not in folder_list:
            folder_list.append(folder_id)
        folders=db.query("""
            select folder_id from project.folder
            where parent_id=%s
        """%folder_id).dictresult()
        forms=db.query("""
            select form_id from project.form where folder_id=%s
        """%folder_id).dictresult()
        for x in forms:
            if x not in form_list:
                form_list.append(x['form_id'])
        if folders!=[]:
            for f in folders:
                folder_list,form_list=getFormsToDelete(f['folder_id'],folder_list,form_list)
        return folder_list,form_list
    except:
        False,False

@bp.route('/getAvailableProjects', methods=['GET','POST'])
@is_logged_in
def getAvailableProjects():
    #obtener los proyectos que están disponibles para clonar
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_projects'})
                if success:
                    if allowed:
                        # projects=db.query("""
                        #     select a.project_id, b.name
                        #     from project.project_users a, project.project b
                        #     where a.user_id = %s
                        #     and a.project_id=b.project_id
                        #     order by b.created desc
                        # """%data['user_id']).dictresult()
                        if data['consultant_mode']==False:
                            projects=db.query("""
                                select distinct(b.project_id), b.name
                                from project.project_users a, project.project b
                                where a.user_id = %s
                                and a.project_id=b.project_id
                                order by b.project_id desc;
                            """%data['user_id']).dictresult()
                        else:
                            projects=db.query("""
                                select distinct(b.project_id), b.name
                                from project.project_users a, project.project b
                                where a.user_id in (select c.user_id from system.user c where c.workspace_id=%s)
                                and a.project_id=b.project_id
                                order by b.project_id desc
                            """%data['workspace_id']).dictresult()

                        response['success']=True
                        response['data']=projects
                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentr obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/saveClonedProject', methods=['GET','POST'])
@is_logged_in
def saveClonedProject():
    #clona un proyecto
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_projects'})
                if success:
                    if allowed:
                        #insertar datos de nuevo proyecto a tabla project.project
                        project_info=copy.deepcopy(data)
                        del project_info['project_id']
                        project_info['created']='now'
                        new_project=db.insert('project.project',project_info)
                        user_1={'user_id':project_info['manager'],'project_id':new_project['project_id']}
                        db.insert('project.project_users',user_1)
                        user_2={'user_id':project_info['partner'],'project_id':new_project['project_id']}
                        db.insert('project.project_users',user_2)
                        #revisa el workspace de quien creó el proyecto para ver que coincida con el resto de los usuarios (cuando se crea un proyecto desde modo consultor)

                        ws=db.query("""
                            select workspace_id from system.user where user_id=%s
                        """%project_info['created_by']).dictresult()
                        if int(ws[0]['workspace_id'])==int(project_info['workspace_id']):
                            if int(project_info['created_by'])!=int(project_info['manager']) and int(project_info['created_by'])!=int(project_info['partner']):
                                user_3={'user_id':project_info['created_by'],'project_id':new_project['project_id']}
                                db.insert('project.project_users',user_3)

                        #insertar carpetas
                        folders=db.query("""
                            select folder_id, name, parent_id
                            from project.folder where project_id=%s order by folder_id asc
                        """%project_info['cloned_project']).dictresult()
                        old_folders_record=[]
                        for f in folders:
                            folder={'name':f['name'],'project_id':new_project['project_id']}
                            if f['parent_id']!=-1:
                                for ofr in old_folders_record:
                                    if int(ofr['old'])==int(f['parent_id']):
                                        folder['parent_id']=ofr['new']
                                        break
                            else:
                                folder['parent_id']=-1
                            inserted_folder=db.insert('project.folder',folder)
                            old_folders_record.append({'old':f['folder_id'],'new':inserted_folder['folder_id']})

                        #insertar formularios
                        old_forms=db.query("""
                            select form_id, project_id, name, columns_number, rows, columns, folder_id
                            from project.form
                            where project_id=%s order by form_id asc
                        """%project_info['cloned_project']).dictresult()

                        old_forms_record=[]
                        for of in old_forms:
                            new_form={
                                'name':of['name'],
                                'columns_number':of['columns_number'],
                                'rows':of['rows'],
                                'columns':of['columns'],
                                'created_by':data['user_id'],
                                'create_date':'now',
                                'status_id':2,
                                'project_id':new_project['project_id']
                            }
                            for ofr in old_folders_record:
                                if int(ofr['old'])==int(of['folder_id']):
                                    new_form['folder_id']=ofr['new']
                                    break
                            inserted_form=db.insert('project.form',new_form)
                            old_forms_record.append({'old':of['form_id'],'new':inserted_form['form_id']})
                            #insertar tabla en esquema form con información del formulario
                            new_table_name='form.project_%s_form_%s'%(new_project['project_id'],inserted_form['form_id'])
                            old_table_name='form.project_%s_form_%s'%(of['project_id'],of['form_id'])
                            db.query("""
                                CREATE TABLE %s as table %s
                            """%(new_table_name,old_table_name))
                            sequence_name='form.project_%s_form_%s_entry_id_seq'%(new_project['project_id'],inserted_form['form_id'])
                            db.query("""
                                CREATE SEQUENCE %s;
                                SELECT setval('%s', coalesce (max(entry_id),0)) FROM %s;
                                ALTER TABLE %s ALTER COLUMN entry_id set default nextval('%s');
                                ALTER TABLE %s ADD PRIMARY KEY (entry_id);
                            """%(sequence_name,sequence_name,new_table_name,new_table_name,sequence_name,new_table_name))
                            exists_rev=db.query("""
                                select column_name from information_schema.columns where table_name='project_%s_form_%s' and column_name='rev_1';
                            """%(new_project['project_id'],inserted_form['form_id'])).dictresult()
                            if exists_rev!=[]:
                                db.query("""
                                    alter table %s drop column rev_1;
                                """%new_table_name)
                            columns_eval=eval(of['columns'])
                            columns_delete=[]
                            for ce in columns_eval:
                                if ce['editable']==True:
                                    columns_delete.append(ce['order'])
                            columns_delete_str=",".join("col_%s=''"%str(e) for e in columns_delete)
                            if columns_delete_str!='':
                                app.logger.info("""
                                    update %s set %s
                                """%(new_table_name,columns_delete_str))
                                db.query("""
                                    update %s set %s
                                """%(new_table_name,columns_delete_str))
                            else:
                                app.logger.info("sin columnas por editables por modificar")
                        response['success']=True
                        response['msg_response']='El proyecto ha sido clonado exitosamente.'

                    else:
                        response['success']=False
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['success']=False
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getSettingsForEditing', methods=['GET','POST'])
@is_logged_in
def getSettingsForEditing():
    #obtiene los datos de la configuración del formulario para poder editarlos
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        allowed_users=db.query("""
                            select created_by as user from project.form where form_id=%s and created_by=%s
                            union
                            select user_id as user from project.form_revisions where form_id=%s and user_id=%s
                            union
                            select manager as user from project.project where project_id=%s and manager=%s
                            union
                            select partner as user from project.project where project_id=%s and partner=%s
                        """%(data['form_id'],data['user_id'],data['form_id'],data['user_id'],data['project_id'],data['user_id'],data['project_id'],data['user_id'])).dictresult()
                        if allowed_users!=[]:
                            settings=db.query("""
                                select b.form_id, b.name, b.status_id,
                                b.columns_number,
                                b.rows,
                                b.columns as columns_info, b.folder_id,
                                (select a.name from project.folder a where a.folder_id=b.folder_id) as folder_name
                                from project.form b
                                where b.form_id=%s
                            """%data['form_id']).dictresult()[0]
                            if int(settings['status_id'])!=7:
                                settings['columns']=eval(settings['columns_info'])
                                del settings['columns_info']
                                response['success']=True
                                response['data']=settings

                            else:
                                response['msg_response']='El formulario ya se encuentra cerrado, por lo tanto no puede ser editado.'
                        else:
                            response['msg_response']='No tienes permisos para editar este formulario.'
                    else:
                        response['msg_response']='No tienes permisos para editar formularios.'
                else:
                    response['msg_response']='Ocurrió un error al intentar validar la información.'
            else:
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getFolderMenuEdit', methods=['GET','POST'])
@is_logged_in
def getFolderMenuEdit():
    #obtener el menú para la edición de configuración de formulario
    #no es necesario validar permisos, porque ya se validó al abrir el modal de edición de formulario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                folders=db.query("""
                    select folder_id, name from project.folder
                    where project_id=%s and parent_id=-1 order by folder_id
                """%data['project_id']).dictresult()
                final_html=""
                if folders!=[]:
                    for f in folders:
                        fh_node='<input type="checkbox" class="form-check-input tree-menu-checkbox-edit folder-checkbox-edit"><li class="selectable-folder-edit"><a href="#" data-folder="%s">%s</a><ul>'%(f['folder_id'],f['name'])
                        sh_node='</ul></li>'
                        res_html=getMenuNodesEdit(f['folder_id'],fh_node,sh_node,data['project_id'])
                        final_html+=res_html
                response['success']=True
                response['menu']='<ul class="file-tree file-tree-edit">%s</ul>'%final_html
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)


def getMenuNodesEdit(folder_id, fh_html,sh_html,project_id):
    #obtiene los nodos del menú
    #no requiere validar permisos
    try:
        folders=db.query("""
            select folder_id, name
            from project.folder
            where parent_id=%s order by folder_id
        """%folder_id).dictresult()
        forms=db.query("""
            select
                form_id, name
            from project.form where project_id=%s and folder_id=%s
            and status_id>2 order by name
        """%(project_id,folder_id)).dictresult()

        node=""

        if folders==[]:
            if forms==[]:
                return fh_html+sh_html
            else:

                forms_str=""
                for f in forms:
                    project_factor=int(project_id)*int(cfg.project_factor)
                    url='/project/%s/%s'%(project_factor,f['form_id'])
                    forms_str+='<li class="selectable-form-edit"><a href="#" id="%s">%s</a></li>'%(f['form_id'],f['name'])
                return fh_html+forms_str+sh_html
        else:
            forms_str=""
            if forms!=[]:
                for f in forms:
                    project_factor=int(project_id)*int(cfg.project_factor)
                    url='/project/%s/%s'%(project_factor,f['form_id'])
                    forms_str+='<li class="selectable-form-edit"><a href="#" id="%s">%s</a></li>'%(f['form_id'],f['name'])
            current_level=""
            for x in folders:
                fh_new_node='<input type="checkbox" class="form-check-input tree-menu-checkbox-edit folder-checkbox-edit"><li class="selectable-folder-edit"><a href="#" data-folder="%s">%s</a><ul>'%(x['folder_id'],x['name'])
                sh_new_node='</ul></li>'
                res_node=getMenuNodesEdit(x['folder_id'],fh_new_node,sh_new_node,project_id)
                current_level+=res_node

            return fh_html+forms_str+current_level+sh_html
    except:
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))

@bp.route('/saveEditFormSettings', methods=['GET','POST'])
@is_logged_in
def saveEditFormSettings():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:

                old_form=db.query("""
                    select * from project.form
                    where form_id=%s
                """%data['form_id']).dictresult()[0]
                old_columns=eval(old_form['columns'])
                new_columns=data['columns_info']
                cols_to_remove=[]
                cols_to_add=[]
                non_originals=[]
                originals=[]
                delete_content=[]

                columns_update=[]
                #recorre info de columnas nuevas
                for nc in new_columns:
                    #se verifica si es columna original
                    if nc['original_%s'%nc['order']]==False:
                        #si no es columna original, será agregada a tabla en esquema form
                        non_originals.append(nc['order'])
                        cols_to_add.append(nc['order'])
                    else:
                        #si es columna original
                        originals.append(nc['order'])
                        #se verifica si hay que borrar contenido de la columna
                        if nc['checkdel_%s'%nc['order']]==True:
                            delete_content.append(nc['order'])
                    #se va armando lista de columnas en el formato en que se va a actualizar a la base de datos
                    columns_update.append({
                        'editable':nc['checkcol_%s'%nc['order']],
                        'order':nc['order'],
                        'name':nc['name']
                    })

                #se recorre configuración antigüa de columnas
                for oc in old_columns:
                    #se verifica si esas columnas se encuentran en las nuevas columnas originales
                    if str(oc['order']) not in originals:
                        #si no se encuentran, se agrega a lista de columnas a eliminar de esquema form
                        cols_to_remove.append(oc['order'])
                #nombre de tabla en esquema form
                form_table="form.project_%s_form_%s"%(data['project_id'],data['form_id'])

                #se actualiza configuración de Columnas
                db.query("""
                    update project.form
                    set name='%s',
                    folder_id=%s,
                    rows=%s,
                    columns_number=%s,
                    columns='%s'
                    where form_id=%s
                """%(data['name'],data['folder_id'],data['rows'],data['columns_number'],str(columns_update).replace("'","''"),data['form_id']))

                #se comprueba si se aumentó el número de filas
                if int(data['rows'])>int(old_form['rows']):
                    rows_to_add=int(data['rows'])-int(old_form['rows'])
                    for i in range(0,rows_to_add):
                        db.insert(form_table,{'form_id':data['form_id'],'project_id':data['project_id']})
                if int(data['rows'])<int(old_form['rows']):
                    rows_to_remove=int(old_form['rows'])-int(data['rows'])
                    db.query("""
                        delete from %s where entry_id in (select entry_id from %s order by entry_id desc limit %s)
                    """%(form_table,form_table,rows_to_remove))


                if cols_to_remove!=[]:
                    for ctr in cols_to_remove:
                        db.query("""
                            alter table %s drop column IF EXISTS %s
                        """%(form_table,'col_%s'%ctr))
                if cols_to_add!=[]:
                    for cta in cols_to_add:
                        db.query("""
                            alter table %s add %s text default ''
                        """%(form_table,'col_%s'%cta))

                if delete_content!=[]:
                    for dc in delete_content:
                        db.query("""
                            update %s
                            set %s=''
                        """%(form_table,'col_%s'%dc))
                response['msg_response']='La configuración ha sido actualizada.'
                response['success']=True
            else:
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/permissionDeleteProject', methods=['GET','POST'])
@is_logged_in
def permissionDeleteProject():
    #verifica si el usuario tiene permisos para eliminar proyectos
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'delete_projects'})
                if success:
                    response['success']=True
                    if allowed:
                        response['allowed']=True
                    else:
                        #revisar si es consultor del espacio de trabajo
                        if data['consultant']==True:
                            consultants=db.query("""
                                select * from system.consultants where user_id=%s
                            """%data['user_id']).dictresult()
                            if consultants!=[]:
                                ws=consultants[0]['workspaces'].split(",")
                                if str(data['workspace_id']) in ws:
                                    response['success']=True
                                    response['allowed']=True
                            else:
                                response['allowed']=False
                                response['msg_response']='No tienes permisos para eliminar proyectos.'
                        else:
                            response['allowed']=False
                            response['msg_response']='No tienes permisos para eliminar proyectos.'
                else:
                    response['msg_response']='Ocurrió un error al intentar procesar la información.'
            else:
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['msg_response']='Ocurrió un error, favor de intenterlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getProjectInfo', methods=['GET','POST'])
@is_logged_in
def getProjectInfo():
    #obtiene los datos del proyecto
    #no requiere verificar permisos
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                project=db.query("""
                    select a.name, a.company_name,
                    to_char(a.start_date,'DD-MM-YYYY') as start_date,
                    to_char(a.finish_date,'DD-MM-YYYY') as finish_date,
                    (select b.name from system.user b where b.user_id=a.manager) as manager,
                    (select b.name from system.user b where b.user_id=a.partner) as partner,
                    a.comments,
                    (select b.name from system.user b where b.user_id=a.created_by) as created_by,
                    to_char(a.created,'DD-MM-YYYY') as created
                    from project.project a
                    where a.project_id=%s
                """%data['project_id']).dictresult()[0]

                info='<p><b>Empresa:</b> {company_name}<br><b>Periodo: </b>{start_date} - {finish_date}<br><b>Creado por: </b>{created_by}<br><b>Fecha de creación: </b>{created}<br><b>Gerente: </b>{manager}<br><b>Socio: </b>{partner}<br><b>Comentarios: </b>{comments}</p>'.format(**project)

                response['success']=True
                response['project_name']=project['name']
                response['project_info']=info

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar procesar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/deleteProject', methods=['GET','POST'])
@is_logged_in
def deleteProject():
    #elimina un proyecto
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'delete_projects'})
                if success:
                    if allowed:
                        forms=db.query("""
                            select form_id from project.form where project_id=%s
                        """%data['project_id']).dictresult()
                        #eliminar tablas de cada formulario (esquema form)
                        for f in forms:
                            table_name="form.project_%s_form_%s"%(data['project_id'],f['form_id'])
                            db.query("""
                                drop table IF EXISTS %s;
                            """%table_name)

                        form_ids=','.join(str(e['form_id']) for e in forms)

                        try:
                            #eliminar comentarios del formulario
                            db.query("""
                                delete from project.form_comments where form_id in (%s)
                            """%form_ids)
                        except:
                            app.logger.info("No  hay comentarios de formularios")

                        #eliminar archivos del proyecto
                        if os.path.exists(os.path.join(cfg.zip_main_folder,'project_%s'%data['project_id'])):
                            shutil.rmtree(os.path.join(cfg.zip_main_folder,'project_%s'%data['project_id']))
                        try:
                            #eliminar registros de archivos de formularios
                            db.query("""
                                delete from project.form_files where form_id in (%s)
                            """%form_ids)
                        except:
                            app.logger.info("No hay archivos de este formulario")

                        try:
                            #eliminar revisores de formularios
                            db.query("""
                                delete from project.form_revisions where form_id in (%s)
                            """%form_ids)
                        except:
                            app.logger.info("No hay registros en form_revisions")

                        try:
                            #eliminar notificaciones
                            db.query("""
                                delete from project.notification where project_id=%s
                            """%data['project_id'])
                        except:
                            app.logger.info("No hay notificaciones del proyecto")

                        try:
                            #elimina carpetas del proyecto
                            db.query("""
                                delete from project.folder where project_id=%s
                            """%data['project_id'])
                        except:
                            app.logger.info("No hay carpetas del proyecto")

                        try:
                            #elimina registros de formularios del proyecto
                            db.query("""
                                delete from project.form where project_id=%s
                            """%data['project_id'])
                        except:
                            app.logger.info("No hay regisros de formulario del proyecto")

                        try:
                            #elimina los usuarios del proyecto
                            db.query("""
                                delete from project.project_users where project_id=%s
                            """%data['project_id'])
                        except:
                            app.logger.info("No hay usuarios del proyecto")

                        try:
                            #elimina registro del proyecto
                            db.query("""
                                delete from project.project where project_id=%s
                            """%data['project_id'])
                        except:
                            app.logger.info("No hay registro del proyecto")

                        response['success']=True
                        response['msg_response']='El proyecto ha sido eliminado.'

                    else:
                        response['msg_response']='No tienes permisos para eliminar proyectos.'
                else:
                    response['msg_response']='Ocurrió un error al intentar procesar la información.'
            else:
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getProjectFormsInfo', methods=['GET','POST'])
@is_logged_in
def getProjectFormsInfo():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                valid_user=db.query("""
                    select manager, partner from project.project where project_id=%s
                """%data['project_id']).dictresult()[0]
                if int(valid_user['manager'])==int(data['user_id']) or int(valid_user['partner'])==int(data['user_id']):
                    forms=db.query("""
                        select a.form_id, a.name, b.status, a.status_id
                        from project.form_status b, project.form a
                        where a.project_id=%s
                        and a.status_id=b.status_id
                    """%data['project_id']).dictresult()
                    for f in forms:
                        if f['status_id']==2:
                            f['link']='/project/%s/createform/step-2/%s'%(int(data['project_id'])*cfg.project_factor,f['form_id'])
                        else:
                            f['link']='/project/%s/%s'%(int(data['project_id'])*cfg.project_factor,f['form_id'])

                    response['data']=forms
                    response['success']=True
                else:
                    response['msg_response']='No tienes permisos para ver esta información.'
            else:
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=True
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/deletePrefilledForm', methods=['GET','POST'])
@is_logged_in
def deletePrefilledForm():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'delete_foldersandforms'})
                if success:
                    if allowed:
                        table_name="form.project_%s_form_%s"%(data['project_id'],data['form_id'])
                        db.query("""
                            drop table IF EXISTS %s;
                        """%table_name)
                        db.query("""
                            delete from project.form where form_id=%s
                        """%data['form_id'])
                        response['success']=True
                    else:
                        response['msg_response']='No tienes permisos para borrar este formulario.'
                else:
                    response['msg_response']='Ocurrió un error al intentar procesar la información.'
            else:
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getProjectEditInfo', methods=['GET','POST'])
@is_logged_in
def getProjectEditInfo():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_projects'})
                if success:
                    response['success']=True
                    if allowed:
                        project=db.query("""
                            select name,company_name,
                            to_char(start_date,'YYYY-MM-DD') as start_date,
                            to_char(finish_date,'YYYY-MM-DD') as finish_date,
                            manager,
                            partner,
                            comments
                            from project.project
                            where project_id=%s and (created_by=%s or manager=%s or partner=%s)
                        """%(data['project_id'],data['user_id'],data['user_id'],data['user_id'])).dictresult()
                        if project!=[]:
                            project_users=db.query("""
                                select a.name, b.user_id
                                from system.user a,
                                project.project_users b
                                where a.user_id=b.user_id
                                and b.project_id=%s
                            """%data['project_id']).dictresult()
                            response['allowed']=True
                            response['project_info']=project[0]
                            response['users']=project_users
                        else:
                            response['msg_response']='No tienes permisos para editar el proyecto.'
                    else:
                        response['msg_response']='No tienes permisos para editar el proyecto.'
                else:
                    response['msg_response']='Ocurrió un error al intentar obtener los datos.'
            else:
                response['msg_response']='Ocurrió un error al intentar procesar los datos.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/saveEditedProjectInfo', methods=['GET','POST'])
@is_logged_in
def saveEditedProjectInfo():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_projects'})
                if success:
                    if allowed:
                        user_allowed=db.query("""
                            select project_id from project.project where project_id=%s
                            and (manager=%s or partner=%s or created_by=%s)
                        """%(data['project_id'],data['user_id'],data['user_id'],data['user_id'])).dictresult()
                        if user_allowed!=[]:
                            for k,v in data.iteritems():
                                if type(v) is not int:
                                    data[k]=v.encode('utf-8')
                            db.query("""
                                update project.project
                                set name='{name}',
                                company_name='{company_name}',
                                start_date='{start_date}',
                                finish_date='{finish_date}',
                                partner={partner},
                                manager={manager},
                                comments='{comments}'
                                where project_id={project_id}
                            """.format(**data))
                            response['success']=True
                            response['msg_response']='Los datos del proyecto han sido actualizados.'
                        else:
                            response['msg_response']='No tienes permisos para modificar los datos de este proyecto.'
                    else:
                        response['msg_response']='No tienes permisos para modificar los datos de este proyecto.'
                else:
                    response['msg_response']='Ocurrió un error al intentar obtener los datos.'
            else:
                response['msg_response']='Ocurrió un error al intentar validar los datos.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getMenuForms', methods=['GET','POST'])
@is_logged_in
def getMenuForms():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        folders=db.query("""
                            select folder_id, name from project.folder
                            where project_id=%s and parent_id=-1 order by folder_id
                        """%data['project_id']).dictresult()
                        final_html=""
                        if folders!=[]:
                            for f in folders:
                                fh_node='<li class="selectable-folder-edit"><a href="#" data-folder="%s">%s</a><ul>'%(f['folder_id'],f['name'])
                                sh_node='</ul></li>'
                                res_html=getMenuFormsNodes(f['folder_id'],fh_node,sh_node,data['project_id'])
                                final_html+=res_html
                        response['success']=True
                        response['menu']='<ul class="file-tree file-tree-edit">%s</ul>'%final_html
                    else:
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['msg_response']='Ocurrió un error al intentar obtener los datos.'
            else:
                response['msg_response']='Ocurrió un error al intentar procesar los datos.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

def getMenuFormsNodes(folder_id, fh_html,sh_html,project_id):
    #obtiene los nodos del menú
    #no requiere validar permisos
    try:
        folders=db.query("""
            select folder_id, name
            from project.folder
            where parent_id=%s order by folder_id
        """%folder_id).dictresult()
        forms=db.query("""
            select
                form_id, name
            from project.form where project_id=%s and folder_id=%s
            and status_id>2 order by name
        """%(project_id,folder_id)).dictresult()

        node=""

        if folders==[]:
            if forms==[]:
                return fh_html+sh_html
            else:
                forms_str=""
                for f in forms:
                    project_factor=int(project_id)*int(cfg.project_factor)
                    url='/project/%s/%s'%(project_factor,f['form_id'])
                    forms_str+='<input type="checkbox" class="form-check-input tree-menu-checkbox-cloned form-checkbox-cloned"><li class="selectable-form-edit"><a href="#" id="%s">%s</a></li>'%(f['form_id'],f['name'])
                return fh_html+forms_str+sh_html
        else:
            forms_str=""
            if forms!=[]:
                for f in forms:
                    project_factor=int(project_id)*int(cfg.project_factor)
                    url='/project/%s/%s'%(project_factor,f['form_id'])
                    forms_str+='<input type="checkbox" class="form-check-input tree-menu-checkbox-cloned form-checkbox-cloned"><li class="selectable-form-edit"><a href="#" id="%s">%s</a></li>'%(f['form_id'],f['name'])
            current_level=""
            for x in folders:
                fh_new_node='<li class="selectable-folder-edit"><a href="#" data-folder="%s">%s</a><ul>'%(x['folder_id'],x['name'])
                sh_new_node='</ul></li>'
                res_node=getMenuFormsNodes(x['folder_id'],fh_new_node,sh_new_node,project_id)
                current_level+=res_node

            return fh_html+forms_str+current_level+sh_html
    except:
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))

@bp.route('/cloneFormAnotherProject', methods=['GET','POST'])
@is_logged_in
def cloneFormAnotherProject():
    response={}
    try:
        response['success']=False
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                success,allowed=GF.checkPermission({'user_id':data['user_id'],'permission':'create_forms'})
                if success:
                    if allowed:
                        form_info=db.query("""
                            select form_id,
                            columns_number, rows,
                            columns from project.form
                            where form_id=%s
                            and project_id=%s
                        """%(data['old_form_id'],data['old_project_id'])).dictresult()[0]
                        old_table_name='form.project_%s_form_%s'%(data['old_project_id'],data['old_form_id'])

                        old_form_content=db.query("""
                            select * from %s order by entry_id asc
                        """%(old_table_name)).dictresult()

                        new_form={
                            'project_id':data['project_id'],
                            'name':data['form_name'].encode('utf-8'),
                            'columns_number':form_info['columns_number'],
                            'rows':form_info['rows'],
                            'columns':form_info['columns'],
                            'created_by':data['user_id'],
                            'create_date':'now',
                            'folder_id':data['folder_id'],
                            'status_id':2,
                            'user_last_update':data['user_id'],
                            'last_updated':'now'
                        }

                        inserted_form=db.insert('project.form',new_form)

                        columns=eval(form_info['columns'])

                        new_table_name='project_%s_form_%s'%(inserted_form['project_id'],inserted_form['form_id'])

                        columns_str=""
                        including_columns=[]
                        for x in columns:
                            columns_str+=" ,col_%s text default ''"%x['order']
                            if x['editable']==False:
                                including_columns.append("col_%s"%x['order'])

                        db.query("""
                            CREATE TABLE form.%s(
                                entry_id serial not null primary key,
                                form_id integer not null,
                                project_id integer not null
                                %s
                            );
                        """%(new_table_name,columns_str))

                        for oi in old_form_content:
                            entry={
                                'form_id':inserted_form['form_id'],
                                'project_id':inserted_form['project_id']
                            }
                            if including_columns!=[]:
                                for ic in including_columns:
                                    entry[ic]=oi[ic]
                            db.insert('form.%s'%new_table_name,entry)

                        response['success']=True
                        response['msg_response']='El formulario ha sido clonado.'

                    else:
                        response['msg_response']='No tienes permisos para realizar esta acción.'
                else:
                    response['msg_response']='Ocurrió un error al intentar obtener los datos.'
            else:
                response['msg_response']='Ocurrió un error al intentar validar los datos.'
        else:
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getConsultantWorkspaces', methods=['GET','POST'])
@is_logged_in
def getConsultantWorkspaces():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                has_ws=db.query("""
                    select * from system.consultants
                    where user_id=%s
                """%data['user_id']).dictresult()
                if has_ws!=[]:
                    workspaces=db.query("""
                        select workspace_id, name from system.workspace
                        where workspace_id in (%s) order by name
                    """%has_ws[0]['workspaces']).dictresult()
                    response['success']=True
                    response['data']=workspaces
                else:
                    response['success']=False
                    response['msg_response']='No existen espacios de trabajo asignados a este consultor.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getConsultantProjects', methods=['GET','POST'])
@is_logged_in
def getConsultantProjects():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                ws_users=db.query("""
                    select user_id from system.user
                    where workspace_id=%s
                """%data['workspace_id']).dictresult()

                if ws_users!=[]:

                    ws_users_str=','.join(str(e['user_id']) for e in ws_users)

                    app.logger.info("""
                        (select a.project_id, a.name, (a.project_id*%d) as project_factor,to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created
                        from project.project a
                        where a.manager in (%s) or a.partner in (%s)
                        union
                        select a.project_id, a.name, (a.project_id*%d) as project_factor,to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created
                        from project.project a, project.project_users b
                        where a.project_id=b.project_id
                        and b.user_id in (%s))
                        order by name asc
                    """%(int(cfg.project_factor),ws_users_str,ws_users_str,int(cfg.project_factor),ws_users_str))

                    projects=db.query("""
                        (select a.project_id, a.name, (a.project_id*%d) as project_factor,to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created
                        from project.project a
                        where a.manager in (%s) or a.partner in (%s)
                        union
                        select a.project_id, a.name, (a.project_id*%d) as project_factor,to_char(a.created,'DD-MM-YYYY HH24:MI:SS') as created
                        from project.project a, project.project_users b
                        where a.project_id=b.project_id
                        and b.user_id in (%s))
                        order by name asc
                    """%(int(cfg.project_factor),ws_users_str,ws_users_str,int(cfg.project_factor),ws_users_str)).dictresult()
                else:
                    projects=[]
                response['data']=projects
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getPublishingInfo', methods=['GET','POST'])
@is_logged_in
def getPublishingInfo():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                is_allowed=db.query("""
                    select edit_publishing_info from system.user where user_id=%s
                """%data['user_id']).dictresult()
                if is_allowed[0]['edit_publishing_info']==True:
                    assigned=db.query("""
                        select assigned_to, project_id, to_char(resolve_before,'YYYY-MM-DD') as resolve_before, notify_assignee, notify_resolved from project.form
                        where form_id=%s
                    """%data['form_id']).dictresult()[0]
                    revisions=db.query("""
                        select user_id, revision_number, currently_assigned, to_char(revision_date,'DD-MM-YYYY'),
                        case when revision_date = '1900-01-01 00:00:00' then 'no' else 'yes' end as already_revised
                        from project.form_revisions
                        where form_id=%s
                        order by revision_number asc
                    """%data['form_id']).dictresult()
                    users=db.query("""
                        select a.user_id, b.name from project.project_users a, system.user b where a.user_id=b.user_id and a.project_id=%s
                    """%assigned['project_id']).dictresult()
                    assigned['revisions']=revisions
                    assigned['users']=users
                    response['data']=assigned
                    response['success']=True

                else:
                    response['success']=False
                    response['msg_response']='No tienes permisos para editar esta información.'
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/editPublishingInfo', methods=['GET','POST'])
@is_logged_in
def editPublishingInfo():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                #comparar para ver si hay cambios
                new_revisions=[]
                new_revisions_wusers=[]

                diff=False #bandera para indicar si hubo algún cambio
                send_to_list=[]
                for k,v in data.iteritems():
                    if k.split("_")[0]=='revision':
                        new_revisions.append(k)
                        new_revisions_wusers.append("%s:%s"%(k,v))

                old_revisions=db.query("""
                    select *, to_char(revision_date,'YYYY-MM-DD') as formatted_date  from project.form_revisions where form_id=%s order by revision_number asc
                """%data['form_id']).dictresult()
                for element in old_revisions:
                    send_to_list.append(element['user_id'])
                form_info=db.query("""
                    select assigned_to, to_char(resolve_before,'YYYY-MM-DD') as resolve_before,
                    notify_assignee, notify_resolved from project.form where form_id=%s
                """%data['form_id']).dictresult()[0]
                send_to_list.append(form_info['assigned_to'])
                send_to_list.append(data['assigned_to'])
                #revisar que si es el mismo número, no sean los mismos revisores
                #revisar si es mismo asignado
                #revisar si es la misma fecha
                notif_changes={'revisor':False,'assigned_to':False,'date':False,'send':False}
                if len(old_revisions)==len(new_revisions):
                    for o in old_revisions:
                        if int(o['user_id'])!=int(data['revision_%s'%o['revision_number']]):
                            diff=True
                            notif_changes['send']=True
                            notif_changes['revisor']=True
                            break
                    if diff==False:
                        if int(data['assigned_to'])!=int(form_info['assigned_to']):
                            diff=True
                            notif_changes['send']=True
                            notif_changes['assigned_to']=True
                        else:
                            if data['resolve_date']!=form_info['resolve_before']:
                                diff=True
                                notif_changes['send']=True
                                notif_changes['date']=True
                            else:
                                if data['notify_resolved']!=form_info['notify_resolved']:
                                    diff=True
                                else:
                                    if data['notify_assignee']!=form_info['notify_assignee']:
                                        diff=True
                                    else:
                                        diff=False
                else:
                    diff=True
                    notif_changes['send']=True
                    notif_changes['revisor']=True

                if diff==True:
                    hist_rev=','.join("revision_%s:%s"%(e['revision_number'],e['user_id']) for e in old_revisions)
                    history={
                        'form_id':data['form_id'],
                        'assigned_to':form_info['assigned_to'],
                        'resolve_before':form_info['resolve_before'],
                        'updated':'now',
                        'updated_by':data['user_id'],
                        'revisors':hist_rev
                    }
                    db.insert("project.publish_history",history)

                    nr=','.join(new_revisions_wusers)

                    db.query("""
                        update project.form
                        set assigned_to=%s,
                        resolve_before='%s',
                        revisions='%s',
                        notify_assignee=%s,
                        notify_resolved=%s
                        where form_id=%s
                    """%(data['assigned_to'],data['resolve_date'],nr,data['notify_assignee'],data['notify_resolved'],data['form_id']))

                    if len(old_revisions)==len(new_revisions):
                        #mismo número de revisores
                        for oldr in old_revisions:
                            if oldr['formatted_date']=='1900-01-01':
                                user=data['revision_%s'%oldr['revision_number']]
                                db.query("""
                                    update project.form_revisions
                                    set user_id=%s,
                                    returned_revision_date='1900-01-01 00:00:00'
                                    where revision_id=%s and form_id=%s
                                """%(user,oldr['revision_id'],data['form_id']))


                    elif len(old_revisions)>len(new_revisions):
                        #si se quitaron revisores
                        db.query("""
                            delete from project.form_revisions
                            where form_id=%s and revision_number>%s
                        """%(data['form_id'],len(new_revisions)))
                        for oldr in old_revisions:
                            if oldr['revision_number']<=len(new_revisions):
                                if oldr['formatted_date']=='1900-01-01':
                                    user=data['revision_%s'%oldr['revision_number']]
                                    db.query("""
                                        update project.form_revisions
                                        set user_id=%s,
                                        returned_revision_date='1900-01-01 00:00:00'
                                        where revision_id=%s and form_id=%s
                                    """%(user,oldr['revision_id'],data['form_id']))
                    else:
                        #si se agregan revisores
                        for oldr in old_revisions:
                            if oldr['formatted_date']=='1900-01-01':
                                user=data['revision_%s'%oldr['revision_number']]
                                db.query("""
                                    update project.form_revisions
                                    set user_id=%s,
                                    returned_revision_date='1900-01-01 00:00:00'
                                    where revision_id=%s and form_id=%s
                                """%(user,oldr['revision_id'],data['form_id']))
                        for x in range(len(old_revisions),len(new_revisions)):
                            revision={
                                'form_id':data['form_id'],
                                'user_id':data['revision_%s'%str(x+1)],
                                'revision_number':x+1,
                                'revision_date':'1900-01-01 00:00:00',
                                'currently_assigned':False
                            }
                            db.insert('project.form_revisions',revision)

                    if notif_changes['send']==True:
                        #obtener información del formulario
                        form_info2=db.query("""
                            select
                                a.name as form_name,
                                a.assigned_to,
                                (select b.name from system.user b where a.assigned_to=b.user_id) as assigned_to_name,
                                to_char(a.resolve_before, 'DD-MM-YYYY') as resolve_before,
                                c.status,
                                a.status_id,
                                a.project_id
                            from project.form a, project.form_status c
                            where a.status_id=c.status_id
                            and a.form_id=%s
                        """%data['form_id']).dictresult()[0]
                        #obtener información de revisores
                        revisors_html=db.query("""
                            select
                                a.name,
                                a.user_id,
                                b.currently_assigned,
                                b.revision_number
                            from
                                system.user a,
                                project.form_revisions b
                            where
                                b.form_id=%s
                            and a.user_id=b.user_id
                            order by b.revision_number
                        """%data['form_id']).dictresult()


                        #obtener correos de destinatarios (asignado y revisores) - incluir nuevos y viejos usuarios asociados al formulario (tanto los que se quitan como los que se agregan)
                        form_info2['top_img']=cfg.img_change_icon
                        form_info2['img_bottom']=cfg.img_rb_logo
                        html='<p style="text-align: center;"><img src="{top_img}" alt="" width="50" height="50" /></p><p><span style="font-family: Verdana, Geneva, sans-serif; color: rgb(93, 93, 93);">Estimado usuario:</span></p><p><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Le notificamos que se realizaron cambios en <strong>{form_name}</strong>:</span></span></p><ul>'
                        html=html.format(**form_info2)
                        link=os.path.join(cfg.host,'project',str(cfg.project_factor*int(form_info2['project_id'])),str(data['form_id']))
                        form_info2['link']='<a href="%s"> aqu&iacute;</a>'%link
                        if notif_changes['assigned_to'] or int(data['assigned_to'])!=int(form_info['assigned_to']):
                            html+='<li><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Usuario asignado</span></span></li>'
                        if notif_changes['revisor']==True:
                            html+='<li><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Revisores</span></span></li>'
                        if notif_changes['date']==True or data['resolve_date']!=form_info['resolve_before']:
                            html+='<li><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Fecha en que deber&aacute; ser resuelto</span></span></li>'

                        html+='</ul><p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Los datos quedan como se muestra a continuaci&oacute;n:</span></span></p><p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Nombre: <em>{form_name}</em></span></span></p><p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Asignado a: <em>{assigned_to_name}</em></span></span></p><p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Resolver antes de: <em>{resolve_before}</em></span></span></p>'
                        html=html.format(**form_info2)
                        for rh in revisors_html:
                            html+='<p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Revisi&oacute;n %s: <em>%s</em></span></span></p>'%(rh['revision_number'],rh['name'])
                            send_to_list.append(rh['user_id'])
                        # revisions_html='<p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Revisi&oacute;n 1:{rev_1}</span></span></p>'

                        html+='<p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Estado actual: <em>{status}</em></span></span></p><p style="line-height: 1;"><span style="color: rgb(93, 93, 93);"><span style="font-family: Verdana, Geneva, sans-serif;">Para acceder, haga click {link}.</span></span></p><p style="line-height: 1;"><br></p><p><img src="{img_bottom}" alt="" width="250" height="70" /></p>'.format(**form_info2)


                        send_list_id=','.join(str(e) for e in send_to_list)
                        sending_emails=db.query("""
                            select distinct email from system.user where user_id in (%s)
                        """%send_list_id).dictresult()
                        list_emails=[]
                        for se in sending_emails:
                            list_emails.append(se['email'])
                        GF.sendMail('Notificación cambios',html,list_emails)

                    response['msg_response']='Los datos han sido actualizados.'
                    #enviar notificaciones
                else:
                    response['msg_response']='Ningún dato fue modificado.'

                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getPublishingHistory', methods=['GET','POST'])
@is_logged_in
def getPublishingHistory():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                history=db.query("""
                    select
                        (select a.name from system.user a where a.user_id=b.assigned_to) as assigned_to,
                        b.revisors,
                        to_char(b.resolve_before,'DD-MM-YYYY') as resolve_before,
                        to_char(b.updated,'DD-MM-YYYY HH:MI:SS') ||' '||(select a.name from system.user a where a.user_id=b.updated_by) as updated
                    from project.publish_history b
                    where form_id=%s
                    order by ph_id desc
                """%data['form_id']).dictresult()

                resp=[]
                for h in history:
                    revisors=h['revisors'].split(",")
                    revs_id=','.join(e.split(":")[1] for e in revisors)
                    revs_names=db.query("""
                        select user_id, name from system.user
                        where user_id in (%s)
                    """%revs_id).dictresult()
                    html='<p><b>Asignado a: </b>{assigned_to}<br><b>Resolver antes de:</b> {resolve_before}<br>'.format(**h)
                    for r in revisors:
                        name=''
                        for rn in revs_names:
                            if int(rn['user_id'])==int(r.split(":")[1]):
                                name=rn['name']
                        html+='<b>Revisor %s:</b> %s<br>'%(r.split(":")[0].split("_")[1],name)
                    resp.append({'html':html,'author':h['updated']})
                response['success']=True
                response['data']=resp

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar obtener los datos, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo más tarde.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        app.logger.info(request.form)
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)
