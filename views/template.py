#--*-- coding: utf-8 --*--
from flask import Flask, render_template, flash, redirect, url_for, session, request, logging, Blueprint, g, send_file
from passlib.hash import sha256_crypt
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
import logging
from .login import is_logged_in
import json
import copy
import sys
import traceback
import os
from openpyxl import Workbook, load_workbook
from .db_connection import getDB
db = getDB()
from flask import current_app as app
import app_config as cfg
from . import general_functions
GF = general_functions.GeneralFunctions()

bp = Blueprint('template', __name__, url_prefix='/templates')


@bp.route('/')
@is_logged_in
def templatesHome():
    #buscar si está en lista de consultores
    is_consultant=db.query("""
        select * from system.consultants
        where user_id=%s
    """%session['user_id']).dictresult()
    if is_consultant!=[]:
        user_info=db.query("""
            select user_id,profile_picture_class,workspace_id
            from system.user where user_id=%s
        """%session['user_id']).dictresult()[0]
        user_info['consultant_workspaces']=is_consultant[0]['workspaces']
        user_info['consultant']=True
        g.user_info=json.dumps(user_info)
        g.profile_picture_class=user_info['profile_picture_class']
        g.notifications=False
        g.consultant=user_info['consultant']
        return render_template('templates.html',g=g)
    else:
        db.query("""
            update system.user_session
            set logged=False,
            finish_session='now'
            where session_id=%s
            and user_id=%s
        """%(session['session_id'],session['user_id']))
        session.clear()
        return redirect(url_for('login.login'))

@bp.route('/preview/<template_factor>/<t_form_id>',methods=['GET','POST'])
@is_logged_in
def resolveForm(template_factor,t_form_id):
    template_id=int(template_factor)/int(cfg.project_factor)
    g=GF.userInfo([{'template_id':template_id},{'template_factor':template_factor},{'t_form_id':t_form_id}])
    g.project_factor=cfg.project_factor
    # has_notifications=db.query("""
    #     select count(*) from project.notification
    #     where project_id=%s and user_to=%s and read=False
    # """%(project_id,session['user_id'])).dictresult()[0]
    # if has_notifications['count']==0:
    #     g.notifications=False
    # else:
    #     g.notifications=True
    return render_template('template_form_preview.html',g=g)

@bp.route('/getTemplatesTable', methods=['GET','POST'])
@is_logged_in
def getTemplatesTable():
    response={}
    try:
        if request.method=='POST':
            templates=db.query("""
                select
                    a.template_id,
                    a.name,
                    a.enabled,
                    case when a.enabled=true then 'Habilitado' else 'Deshabilitado' end as enabled_name,
                    to_char(a.last_updated,'DD-MM-YYYY HH:MI:SS') as last_updated,
                    (select count(b.t_form_id) from templates.t_forms b where b.template_id=a.template_id) as forms_total
                from
                    templates.templates a
                order by a.name asc
                offset %s limit %s
            """%(int(request.form['start']),int(request.form['length']))).dictresult()
            for x in templates:
                x['actions']='<div class="row row-wo-margin" data-templateid="%s"><a href="#mod_edit_template" class="template-edit-a"  data-toggle="modal"><i class="fa fa-edit"></i></a><a href="#mod_create_form_for_template" class="template-newform-a" data-toggle="modal"><i class="fa fa-file-text-o"></i></a><a href="#" class="template-delete-a"><i class="fa fa-trash-o"></i></a></div>'%(str(x['template_id']))

            templates_total=db.query("""
                select count(*) from templates.templates
            """).dictresult()

            response['data']=templates
            response['recordsTotal']=templates_total[0]['count']
            response['recordsFiltered']=templates_total[0]['count']
            response['success']=True
        else:
            response['suceess']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info= sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/saveNewTemplate', methods=['GET','POST'])
@is_logged_in
def saveNewTemplate():
    #obtener los proyectos a los que se encuentra agregado el usuario
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                db.insert("templates.templates",data)
                response['success']=True
                response['msg_response']='La plantilla fue agregada correctamente.'
            else:
                response['suceess']=False
                response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error al intentar procesar la información.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info= sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getFirstTempMenuFolders', methods=['GET','POST'])
@is_logged_in
def getFirstTempMenuFolders():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                folders=db.query("""
                    select
                        t_folder_id,
                        name,
                        template_id
                    from
                        templates.t_folders
                    where
                        template_id=%s
                        and parent_id=-1
                    order by t_folder_id asc
                """%data['template_id']).dictresult()

                div_class='folder-icon-div'
                check_class='checkbox-folder-menu'

                html=''
                for x in folders:
                    html+='<div class="'+div_class+'"><input type="checkbox" class="'+check_class+'" data-document="%s"><div style="display:grid;"><a data-toggle="tooltip" title="%s" class="mp-a-folder"><i class="icon-folder-menu"></i></a><span class="block-with-text" data-toggle="tooltip" title="%s">%s</span></div></div>'%(x['t_folder_id'],x['name'].decode('utf8'),x['name'].decode('utf8'),x['name'].decode('utf8'))

                response['data']=html


                # response['data']=folders
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        app.logger.info(request.form)
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)


@bp.route('/getTempSubfoldersForms', methods=['GET','POST'])
@is_logged_in
def getTempSubfoldersForms():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                folders=db.query("""
                    select
                        t_folder_id,
                        name,
                        template_id
                    from
                        templates.t_folders
                    where
                        template_id=%s
                        and parent_id=%s
                    order by t_folder_id asc
                """%(data['template_id'],data['t_folder_id'])).dictresult()

                forms=db.query("""
                    select
                        t_form_id,
                        name,
                        template_id
                    from
                        templates.t_forms
                    where
                        t_folder_id=%s
                    and template_id=%s
                """%(data['t_folder_id'],data['template_id'])).dictresult()


                div_class='folder-icon-div'
                form_div_class='form-icon-div'
                check_class='checkbox-folder-menu'
                check_form_class='checkbox-form-menu'


                html=''
                for x in folders:
                    html+='<div class="'+div_class+'"><input type="checkbox" class="'+check_class+'" data-document="%s"><div style="display:grid;"><a data-toggle="tooltip" title="%s" class="mp-a-folder"><i class="icon-folder-menu"></i></a><span class="block-with-text" data-toggle="tooltip" title="%s">%s</span></div></div>'%(x['t_folder_id'],x['name'].decode('utf8'),x['name'].decode('utf8'),x['name'].decode('utf8'))

                for y in forms:
                    link=os.path.join(cfg.host,'templates','preview',str(cfg.project_factor*int(y['template_id'])),str(y['t_form_id']))
                    icon_class='icon-form-menu'




                    html+='<div class="'+form_div_class+'"><input type="checkbox" class="'+check_form_class+'" data-document="%s"><div style="display:grid;"><a data-toggle="tooltip" title="%s" class="mp-a-folder" href="%s" target="_blank"><i class="%s"></i></a><span class="block-with-text" data-toggle="tooltip" title="%s">%s</span></div></div>'%(y['t_form_id'],y['name'].decode('utf8'),link,icon_class,y['name'].decode('utf8'),y['name'].decode('utf8'))

                response['data']=html
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        app.logger.info(request.form)
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)


@bp.route('/returnTempSubFolder', methods=['GET','POST'])
@is_logged_in
def returnTempSubFolder():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                parent=db.query("""
                    select
                    parent_id from templates.t_folders
                    where t_folder_id=%s
                    and template_id=%s
                """%(data['parent_id'],data['template_id'])).dictresult()[0]


                folders=db.query("""
                    select
                        t_folder_id,
                        name,
                        template_id
                    from
                        templates.t_folders
                    where
                        template_id=%s
                        and parent_id=%s
                    order by t_folder_id asc
                """%(data['template_id'],parent['parent_id'])).dictresult()

                forms=db.query("""
                    select
                        t_form_id,
                        name,
                        template_id
                    from
                        templates.t_forms
                    where
                        t_folder_id=%s
                    and template_id=%s
                """%(parent['parent_id'],data['template_id'])).dictresult()


                div_class='folder-icon-div'
                form_div_class='form-icon-div'
                check_class='checkbox-folder-menu'
                check_form_class='checkbox-form-menu'


                html=''
                for x in folders:

                    html+='<div class="'+div_class+'"><input type="checkbox" class="'+check_class+'" data-document="%s"><div style="display:grid;"><a data-toggle="tooltip" title="%s" class="mp-a-folder"><i class="icon-folder-menu"></i></a><span class="block-with-text" data-toggle="tooltip" title="%s">%s</span></div></div>'%(x['t_folder_id'],x['name'].decode('utf8'),x['name'].decode('utf8'),x['name'].decode('utf8'))

                for y in forms:

                    icon_class='icon-form-menu'
                    link=os.path.join(cfg.host,'templates','preview',str(cfg.project_factor*int(y['template_id'])),str(y['t_form_id']))


                    html+='<div class="'+form_div_class+'"><input type="checkbox" class="'+check_form_class+'" data-document="%s"><div style="display:grid;"><a data-toggle="tooltip" title="%s" class="mp-a-folder" href="%s" target="_blank"><i class="%s"></i></a><span class="block-with-text" data-toggle="tooltip" title="%s">%s</span></div></div>'%(y['t_form_id'],y['name'].decode('utf8'),link,icon_class,y['name'].decode('utf8'),y['name'].decode('utf8'))

                response['data']=html
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        app.logger.info(request.form)
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/saveTempFolder', methods=['GET','POST'])
@is_logged_in
def saveTempFolder():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                if data['mode']=='new':
                    db.insert('templates.t_folders',data)
                    response['msg_response']='La carpeta ha sido agregada.'
                else:
                    db.query("""
                        update templates.t_folders
                        set name='%s' where t_folder_id=%s
                    """%(data['name'],data['t_folder_id']))
                    response['msg_response']='La carpeta ha sido actualizada.'
                response['success']=True

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        app.logger.info(request.form)
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/saveTempForm', methods=['GET','POST'])
@is_logged_in
def saveTempForm():
    response={}
    try:
        data=request.form.to_dict()


        files=request.files
        file_path=cfg.uploaded_forms_files_path
        file=files[data['file_name']]
        filename = secure_filename(file.filename)
        file.save(os.path.join(file_path, filename))
        read_file=load_workbook(os.path.join(file_path,filename))
        ws = read_file.worksheets[0]
        # ws=read_file['Sheet1']

        form={
            'template_id':data['template_id'],
            'name':data['name'],
            'created_by':int(data['user_id']),
            'create_date':'now',
            't_folder_id':int(data['t_folder_id']),
            'columns_number':int(ws.max_column),
            'rows':int(ws.max_row)
        }

        app.logger.info(ws.max_column)
        columns=[]
        for x in range(1,int(ws.max_column)+1):
            column={
                'order':x,
                'name':ws.cell(row=1, column=x).value
            }
            if ws.cell(row=2, column=x).value is None:
                column['editable']=True
            else:
                column['editable']=False
            columns.append(column)
        form['columns']=str(columns)
        db.insert('templates.t_forms',form)

        table_name='template_%s_tform_%s'%(form['template_id'],form['t_form_id'])


        columns_str=""
        for c in columns:
            columns_str+=" ,col_%s text default ''"%c['order']
        db.query("""
            CREATE TABLE template_tables.%s(
                t_entry_id serial not null primary key,
                t_form_id integer not null,
                template_id integer not null
                %s
            );
        """%(table_name,columns_str))

        for i in range(0,int(form['rows'])-1):
            ins={
                't_form_id':form['t_form_id'],
                'template_id':form['template_id']
            }
            for c in columns:
                if c['editable']==False:
                    if ws.cell(row=i+2, column=c['order']).value is not None:
                        if type(ws.cell(row=i+2, column=c['order']).value)=='unicode':
                            ins['col_%s'%c['order']]=(ws.cell(row=i+2, column=c['order']).value).encode('utf-8')
                        else:
                            ins['col_%s'%c['order']]=ws.cell(row=i+2, column=c['order']).value
                    else:
                        ins['col_%s'%c['order']]=''
            db.insert('template_tables.%s'%table_name,ins)
        response['success']=True
        response['msg_response']='El formulario ha sido agregado.'
        response['t_form_id']=form['t_form_id']
        response['template_factor']=int(cfg.project_factor)*int(data['template_id'])

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        GF.sendErrorMail(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/showTempFormTable', methods=['GET','POST'])
@is_logged_in
def showTempFormTable():
    #obtiene la tabla del formulario para su configuración
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:

                page=(int(data['page'])*10)-10
                form_info=db.query("""
                    select
                        t_form_id,
                        template_id,
                        columns,
                        name
                    from templates.t_forms
                    where t_form_id=%s
                """%(data['t_form_id'])).dictresult()
                columns=eval(form_info[0]['columns'])
                #check if table exists
                table_name='template_%s_tform_%s'%(form_info[0]['template_id'],form_info[0]['t_form_id'])



                table_str='<table class="table table-bordered table-responsive-md table-striped text-center" id="grdPrefilledForm"><thead><tr>'
                for c in columns:
                    table_str+='<th class="text-center">%s</th>'%c['name']
                table_str+='</tr></thead><tbody>'

                table_info=db.query("""
                    select * from template_tables.%s order by t_entry_id
                    offset %s limit 10
                """%(table_name,page)).dictresult()

                total_table_info=db.query("""
                    select count(*) from template_tables.%s
                """%table_name).dictresult()[0]['count']


                for t in table_info:
                    keys=sorted(t.iteritems())
                    table_str+='<tr>'
                    for k in keys:
                        if k[0].split('_')[0]=='col':
                            table_str+='<td class="pt-3-half" contenteditable="false" name="%s" data-entry="%s">%s</td>'%(k[0],t['t_entry_id'],k[1].decode('utf8'))
                        else:
                            table_str+='</tr>'
                            break
                table_str+='</tbody></table>'

                if int(total_table_info)%10!=0:
                    num_buttons=int(int(total_table_info)/10)+1
                else:
                    num_buttons=int(int(total_table_info)/10)
                buttons='<div class="btn-group btn-group-sm" role="group"><input class="paging-toolbar-number" type="text" readonly id="temp_paging_toolbar_number"/>'
                for b in range (0,num_buttons):
                    buttons+='<button type="button" class="btn btn-secondary form-paging-toolbar" data-number="%s">%s</button>'%(int(b+1),int(b+1))
                buttons+='</div>'

                response['form_name']=form_info[0]['name']
                response['success']=True
                response['html']=table_str
                response['paging_toolbar']=buttons

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error la intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/getTemplateEditInfo', methods=['GET','POST'])
@is_logged_in
def getTemplateEditInfo():
    #obtiene la tabla del formulario para su configuración
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                template_info=db.query("""
                    select name, template_id, enabled, case when enabled=True then 'Habilitado' else 'Deshabilitado' end as enabled_text
                    from templates.templates
                    where template_id=%s
                """%data['template_id']).dictresult()[0]

                template_ws=db.query("""
                    select a.workspace_id,
                    'true' as checked,
                    (select b.name from system.workspace b where b.workspace_id=a.workspace_id) as workspace_name
                    from templates.templates_workspace a where a.template_id=%s
                """%data['template_id']).dictresult()

                filter=""
                if template_ws!=[]:

                    checked_ws=','.join(str(e['workspace_id']) for e in template_ws)
                    filter="and workspace_id not in (%s)"%checked_ws

                non_ch_ws=db.query("""
                    select workspace_id, 'false' as checked, name as workspace_name
                    from system.workspace where workspace_id<>-1  %s
                """%filter).dictresult()


                template_info['workspaces']=template_ws+non_ch_ws
                response['data']=template_info
                response['success']=True


            else:
                response['success']=False
                response['msg_response']='Ocurrió un error la intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/saveEditedTemp', methods=['GET','POST'])
@is_logged_in
def saveEditedTemp():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                for k,v in data.iteritems():
                    if k.split("_")[0]=="ws":
                        exists=db.query("""
                            select * from templates.templates_workspace
                            where template_id=%s and workspace_id=%s
                        """%(data['template_id'],k.split("_")[1])).dictresult()
                        if exists==[]:
                            if v==True:
                                db.insert("templates.templates_workspace",{'template_id':data['template_id'],'workspace_id':k.split("_")[1]})
                        else:
                            if v==False:
                                db.query("""
                                    delete from templates.templates_workspace
                                    where template_id=%s and workspace_id=%s
                                """%(data['template_id'],k.split("_")[1]))
                db.query("""
                    update templates.templates
                    set name='%s',
                    enabled=%s,
                    last_updated='now'
                    where template_id=%s
                """%(data['name'],data['enabled'],data['template_id']))

                response['success']=True
                response['msg_response']='La plantilla ha sido actualizada.'

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error la intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)

@bp.route('/deleteTemplate', methods=['GET','POST'])
@is_logged_in
def deleteTemplate():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                db.query("""
                    delete from templates.templates_workspace
                    where template_id=%s
                """%data['template_id'])

                forms=db.query("""
                    select t_form_id from templates.t_forms
                    where template_id=%s
                """%data['template_id']).dictresult()

                if forms!=[]:
                    for x in forms:
                        db.query("""
                            drop table template_tables.template_%s_tform_%s
                        """%(data['template_id'],x['t_form_id']))
                db.query("""
                    delete from templates.t_forms where template_id=%s
                """%data['template_id'])

                db.query("""
                    delete from templates.t_folders where template_id=%s
                """%data['template_id'])

                db.query("""
                    delete from templates.templates where template_id=%s
                """%data['template_id'])

                response['success']=True
                response['msg_response']='La plantilla ha sido eliminada.'

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error la intentar obtener la información, favor de intentarlo de nuevo.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        exc_info=sys.exc_info()
        app.logger.info(traceback.format_exc(exc_info))
        GF.sendErrorMail(traceback.format_exc(exc_info))
    return json.dumps(response)
